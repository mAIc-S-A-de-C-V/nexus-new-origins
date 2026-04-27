import time
import asyncio
import csv
import io
import json
import os
from typing import Optional, Any
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, Header, Query, Depends, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_
import httpx
from models import (
    ConnectorConfig, ConnectorCreateRequest,
    ConnectorUpdateRequest, ConnectionTestResult, ConnectorPublicView
)
from database import ConnectorRow, get_session
from schema_fetcher import fetch_schema, test_credentials
from credential_crypto import encrypt_credentials, decrypt_credentials
from db_connector import (
    list_tables as db_list_tables,
    table_schema as db_table_schema,
    preview_table as db_preview_table,
    run_query as db_run_query,
    test_db_connection,
    _build_db_config,
)

router = APIRouter()

EVENT_LOG_URL = os.environ.get("EVENT_LOG_SERVICE_URL", "http://event-log-service:8005")
ONTOLOGY_API = os.environ.get("ONTOLOGY_SERVICE_URL", "http://ontology-service:8004")
WHATSAPP_API = os.environ.get("WHATSAPP_SERVICE_URL", "http://whatsapp-service:8025")

# In-memory staging area for file upload rows, keyed by connector_id
_file_staging: dict[str, dict[str, Any]] = {}

DB_CONNECTOR_TYPES = {"POSTGRESQL", "MYSQL"}


async def _emit_event(payload: dict) -> None:
    """Fire-and-forget event emission to event-log-service."""
    try:
        async with httpx.AsyncClient(timeout=3.0) as client:
            await client.post(f"{EVENT_LOG_URL}/events", json=payload)
    except Exception:
        pass  # non-critical — never fail a request because of event emission


def _row_to_model(row: ConnectorRow) -> ConnectorConfig:
    return ConnectorConfig(
        id=row.id,
        name=row.name,
        type=row.type,
        category=row.category,
        status=row.status,
        description=row.description,
        base_url=row.base_url,
        auth_type=row.auth_type,
        credentials=decrypt_credentials(row.credentials),
        headers=row.headers,
        pagination_strategy=row.pagination_strategy,
        active_pipeline_count=row.active_pipeline_count,
        last_sync=row.last_sync,
        last_sync_row_count=row.last_sync_row_count,
        schema_hash=row.schema_hash,
        tags=row.tags or [],
        config=row.config,
        created_at=row.created_at,
        updated_at=row.updated_at,
        tenant_id=row.tenant_id,
        visibility=row.visibility or "tenant",
        created_by=row.created_by,
    )


@router.get("", response_model=list[ConnectorPublicView])
async def list_connectors(
    category: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    x_tenant_id: Optional[str] = Header(None),
    x_user_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    tenant_id = x_tenant_id or "tenant-001"
    stmt = select(ConnectorRow).where(ConnectorRow.tenant_id == tenant_id)
    # Visibility filter: show tenant-visible connectors + private ones owned by this user
    if x_user_id:
        stmt = stmt.where(
            or_(
                ConnectorRow.visibility == "tenant",
                ConnectorRow.visibility.is_(None),
                (ConnectorRow.visibility == "private") & (ConnectorRow.created_by == x_user_id),
            )
        )
    else:
        stmt = stmt.where(or_(ConnectorRow.visibility == "tenant", ConnectorRow.visibility.is_(None)))
    if category:
        stmt = stmt.where(ConnectorRow.category == category)
    if status:
        stmt = stmt.where(ConnectorRow.status == status)
    result = await db.execute(stmt)
    return [ConnectorPublicView.from_config(_row_to_model(r)) for r in result.scalars().all()]


@router.post("", response_model=ConnectorPublicView, status_code=201)
async def create_connector(
    req: ConnectorCreateRequest,
    x_tenant_id: Optional[str] = Header(None),
    x_user_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    tenant_id = x_tenant_id or "tenant-001"
    from uuid import uuid4
    row = ConnectorRow(
        id=str(uuid4()),
        tenant_id=tenant_id,
        name=req.name,
        type=req.type,
        category=req.category,
        status="idle",
        description=req.description,
        base_url=req.base_url,
        auth_type=req.auth_type,
        credentials=encrypt_credentials(req.credentials),
        headers=req.headers,
        pagination_strategy=req.pagination_strategy,
        tags=req.tags,
        config=req.config,
        visibility=req.visibility or "tenant",
        created_by=x_user_id,
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return ConnectorPublicView.from_config(_row_to_model(row))


INTERNAL_SECRET = os.environ.get("INTERNAL_SECRET", "nexus-internal")


@router.get("/{connector_id}/internal")
async def get_connector_internal(
    connector_id: str,
    x_tenant_id: Optional[str] = Header(None),
    x_internal: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    """Service-to-service: returns the full connector including decrypted credentials.
    Guarded by x-internal header. Only for pipeline-service / agent-service / etc."""
    if x_internal != INTERNAL_SECRET:
        raise HTTPException(status_code=403, detail="Internal endpoint — missing x-internal header")
    tenant_id = x_tenant_id or "tenant-001"
    result = await db.execute(
        select(ConnectorRow).where(
            ConnectorRow.id == connector_id,
            ConnectorRow.tenant_id == tenant_id,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Connector not found")
    model = _row_to_model(row)
    return model.model_dump()


@router.get("/{connector_id}", response_model=ConnectorPublicView)
async def get_connector(
    connector_id: str,
    x_tenant_id: Optional[str] = Header(None),
    x_user_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    tenant_id = x_tenant_id or "tenant-001"
    stmt = select(ConnectorRow).where(
        ConnectorRow.id == connector_id,
        ConnectorRow.tenant_id == tenant_id,
    )
    # Visibility filter: only allow access to private connectors by their owner
    if x_user_id:
        stmt = stmt.where(
            or_(
                ConnectorRow.visibility == "tenant",
                ConnectorRow.visibility.is_(None),
                (ConnectorRow.visibility == "private") & (ConnectorRow.created_by == x_user_id),
            )
        )
    else:
        stmt = stmt.where(or_(ConnectorRow.visibility == "tenant", ConnectorRow.visibility.is_(None)))
    result = await db.execute(stmt)
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Connector not found")
    return ConnectorPublicView.from_config(_row_to_model(row))


@router.put("/{connector_id}", response_model=ConnectorPublicView)
async def update_connector(
    connector_id: str,
    req: ConnectorUpdateRequest,
    x_tenant_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    tenant_id = x_tenant_id or "tenant-001"
    result = await db.execute(
        select(ConnectorRow).where(
            ConnectorRow.id == connector_id,
            ConnectorRow.tenant_id == tenant_id,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Connector not found")

    update_data = req.model_dump(exclude_none=True)
    for key, value in update_data.items():
        if key == "credentials":
            setattr(row, key, encrypt_credentials(value))
        else:
            setattr(row, key, value)
    row.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(row)
    return ConnectorPublicView.from_config(_row_to_model(row))


@router.delete("/{connector_id}", status_code=204)
async def delete_connector(
    connector_id: str,
    x_tenant_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    tenant_id = x_tenant_id or "tenant-001"
    result = await db.execute(
        select(ConnectorRow).where(
            ConnectorRow.id == connector_id,
            ConnectorRow.tenant_id == tenant_id,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Connector not found")
    await db.delete(row)
    await db.commit()


@router.post("/{connector_id}/test", response_model=ConnectionTestResult)
async def test_connection(
    connector_id: str,
    x_tenant_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    tenant_id = x_tenant_id or "tenant-001"
    result = await db.execute(
        select(ConnectorRow).where(
            ConnectorRow.id == connector_id,
            ConnectorRow.tenant_id == tenant_id,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Connector not found")

    test_cfg = dict(row.config or {})
    test_cfg.setdefault("connector_id", str(connector_id))

    success, message, latency_ms = await test_credentials(
        connector_type=row.type,
        base_url=row.base_url,
        credentials=decrypt_credentials(row.credentials),
        config=test_cfg,
        db=db,
    )

    # Update connector status based on test result
    row.status = "active" if success else "error"
    row.updated_at = datetime.now(timezone.utc)
    await db.commit()

    from uuid import uuid4
    asyncio.create_task(_emit_event({
        "id": str(uuid4()),
        "case_id": connector_id,
        "activity": "CONNECTOR_TEST_PASSED" if success else "CONNECTOR_TEST_FAILED",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "object_type_id": "",
        "object_id": connector_id,
        "pipeline_id": "",
        "connector_id": connector_id,
        "tenant_id": tenant_id,
        "attributes": {"latency_ms": latency_ms, "message": message},
    }))

    return ConnectionTestResult(
        success=success,
        latency_ms=latency_ms,
        message=message,
        error=None if success else message,
    )


@router.get("/{connector_id}/schema")
async def get_schema(
    connector_id: str,
    x_tenant_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    tenant_id = x_tenant_id or "tenant-001"
    result = await db.execute(
        select(ConnectorRow).where(
            ConnectorRow.id == connector_id,
            ConnectorRow.tenant_id == tenant_id,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Connector not found")

    cfg = dict(row.config or {})
    # Inject connector_id so type-specific fetchers (e.g. WHATSAPP) can use it
    cfg.setdefault("connector_id", str(connector_id))

    raw_schema, sample_rows, error = await fetch_schema(
        connector_type=row.type,
        base_url=row.base_url,
        credentials=decrypt_credentials(row.credentials),
        config=cfg,
        db=db,
        last_sync=row.last_sync,
    )

    if not error:
        from uuid import uuid4
        asyncio.create_task(_emit_event({
            "id": str(uuid4()),
            "case_id": connector_id,
            "activity": "CONNECTOR_SCHEMA_FETCHED",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "object_type_id": "",
            "object_id": connector_id,
            "pipeline_id": "",
            "connector_id": connector_id,
            "tenant_id": tenant_id,
            "attributes": {"row_count": len(sample_rows) if sample_rows else 0},
        }))

    return {
        "connector_id": connector_id,
        "connector_type": row.type,
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "schema": raw_schema,
        "sample_rows": sample_rows,
        "error": error,
    }


@router.post("/{connector_id}/fetch-row")
async def fetch_row(
    connector_id: str,
    body: dict,
    x_tenant_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    """
    Call this connector's configured endpoint with query param overrides.
    Used by the pipeline ENRICH node to perform per-row detail lookups.
    Body: {"params": {"id": "INC-123"}}
    Returns: {"row": {...}, "rows": [...]}
    """
    tenant_id = x_tenant_id or "tenant-001"
    result = await db.execute(
        select(ConnectorRow).where(
            ConnectorRow.id == connector_id,
            ConnectorRow.tenant_id == tenant_id,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Connector not found")

    overrides: dict = body.get("params", {})
    config = dict(row.config or {})
    existing_qp = dict(config.get("queryParams") or {})
    existing_qp.update({k: str(v) for k, v in overrides.items()})
    config["queryParams"] = existing_qp

    _, sample_rows, error = await fetch_schema(
        connector_type=row.type,
        base_url=row.base_url,
        credentials=decrypt_credentials(row.credentials),
        config=config,
        db=db,
    )

    if error:
        raise HTTPException(status_code=502, detail=error)

    return {
        "row": sample_rows[0] if sample_rows else {},
        "rows": sample_rows,
    }


@router.patch("/{connector_id}/last-sync")
async def update_last_sync(
    connector_id: str,
    x_tenant_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    """Mark the connector as synced right now (called by pipeline executor after a successful run)."""
    tenant_id = x_tenant_id or "tenant-001"
    result = await db.execute(
        select(ConnectorRow).where(
            ConnectorRow.id == connector_id,
            ConnectorRow.tenant_id == tenant_id,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Connector not found")
    row.last_sync = datetime.now(timezone.utc)
    await db.commit()
    return {"ok": True, "last_sync": row.last_sync.isoformat()}


@router.put("/{connector_id}/inference")
async def save_inference_result(
    connector_id: str,
    body: dict,
    x_tenant_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    """Persist an inference result for a connector (called after Claude inference completes)."""
    tenant_id = x_tenant_id or "tenant-001"
    result = await db.execute(
        select(ConnectorRow).where(
            ConnectorRow.id == connector_id,
            ConnectorRow.tenant_id == tenant_id,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Connector not found")

    row.inference_result = body
    row.inference_ran_at = datetime.now(timezone.utc)
    await db.commit()
    return {"status": "saved"}


@router.get("/{connector_id}/inference")
async def get_inference_result(
    connector_id: str,
    x_tenant_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    """Retrieve the persisted inference result for a connector."""
    tenant_id = x_tenant_id or "tenant-001"
    result = await db.execute(
        select(ConnectorRow).where(
            ConnectorRow.id == connector_id,
            ConnectorRow.tenant_id == tenant_id,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Connector not found")
    if not row.inference_result:
        raise HTTPException(status_code=404, detail="No inference result found for this connector")
    return {
        "connector_id": connector_id,
        "inference_result": row.inference_result,
        "inference_ran_at": row.inference_ran_at.isoformat() if row.inference_ran_at else None,
    }


# ── File Upload Endpoints ─────────────────────────────────────────────────────


def _detect_type(values: list) -> str:
    """Heuristic type detection from a list of sample values."""
    non_null = [v for v in values if v is not None and str(v).strip() != ""]
    if not non_null:
        return "string"
    for v in non_null:
        try:
            int(v)
            continue
        except (ValueError, TypeError):
            break
    else:
        return "integer"
    for v in non_null:
        try:
            float(v)
            continue
        except (ValueError, TypeError):
            break
    else:
        return "float"
    # Check for boolean-like
    bool_vals = {"true", "false", "yes", "no", "1", "0"}
    if all(str(v).strip().lower() in bool_vals for v in non_null):
        return "boolean"
    return "string"


def _parse_csv(content: bytes) -> list[dict]:
    """Parse CSV content with auto-detected delimiter."""
    text = content.decode("utf-8-sig")
    # Auto-detect delimiter
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = None
    reader = csv.DictReader(io.StringIO(text), dialect=dialect) if dialect else csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_excel(content: bytes) -> list[dict]:
    """Parse Excel .xlsx first sheet using openpyxl."""
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(next(rows_iter, []))]
    if not headers:
        return []
    result = []
    for row in rows_iter:
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = val
        result.append(record)
    wb.close()
    return result


def _parse_json(content: bytes) -> list[dict]:
    """Parse JSON as array of objects."""
    data = json.loads(content)
    if isinstance(data, list):
        return [r for r in data if isinstance(r, dict)]
    if isinstance(data, dict):
        # Try common wrapper keys
        for key in ("data", "results", "items", "records", "rows"):
            if isinstance(data.get(key), list):
                return [r for r in data[key] if isinstance(r, dict)]
        # Single object
        return [data]
    return []


@router.post("/{connector_id}/upload")
async def upload_file(
    connector_id: str,
    file: UploadFile = File(...),
    x_tenant_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    """
    Accept a CSV, Excel, or JSON file upload, parse it, stage the rows,
    and return column metadata with sample values.
    """
    tenant_id = x_tenant_id or "tenant-001"
    result = await db.execute(
        select(ConnectorRow).where(
            ConnectorRow.id == connector_id,
            ConnectorRow.tenant_id == tenant_id,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Connector not found")

    content = await file.read()
    filename = file.filename or "unknown"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    content_type = file.content_type or ""

    try:
        if ext == "csv" or "csv" in content_type:
            rows = _parse_csv(content)
        elif ext in ("xlsx", "xls") or "spreadsheet" in content_type or "excel" in content_type:
            rows = _parse_excel(content)
        elif ext == "json" or "json" in content_type:
            rows = _parse_json(content)
        else:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type: .{ext}. Accepted: CSV, XLSX, JSON",
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    if not rows:
        raise HTTPException(status_code=400, detail="File contains no parseable rows")

    # Stage the rows in memory
    _file_staging[connector_id] = {
        "rows": rows,
        "file_name": filename,
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
    }

    # Build column metadata
    all_keys = list(dict.fromkeys(k for r in rows for k in r.keys()))
    columns = []
    for col_name in all_keys:
        sample_vals = [r.get(col_name) for r in rows[:3]]
        all_vals = [r.get(col_name) for r in rows[:50]]
        columns.append({
            "name": col_name,
            "detected_type": _detect_type(all_vals),
            "sample_values": sample_vals,
        })

    return {
        "columns": columns,
        "row_count": len(rows),
        "file_name": filename,
    }


@router.post("/{connector_id}/import")
async def import_file(
    connector_id: str,
    body: dict,
    x_tenant_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    """
    Import staged file rows into the ontology service.
    Body: { object_type_id, pk_field, field_mappings: {source: target} }
    """
    tenant_id = x_tenant_id or "tenant-001"
    result = await db.execute(
        select(ConnectorRow).where(
            ConnectorRow.id == connector_id,
            ConnectorRow.tenant_id == tenant_id,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Connector not found")

    staged = _file_staging.get(connector_id)
    if not staged:
        raise HTTPException(status_code=400, detail="No staged file data. Upload a file first.")

    object_type_id = body.get("object_type_id")
    pk_field = body.get("pk_field", "id")
    field_mappings: dict = body.get("field_mappings", {})

    if not object_type_id:
        raise HTTPException(status_code=400, detail="object_type_id is required")

    # Apply field mappings to staged rows
    mapped_rows = []
    for src_row in staged["rows"]:
        mapped = {}
        for src_key, tgt_key in field_mappings.items():
            if src_key in src_row:
                mapped[tgt_key] = src_row[src_key]
        # If no mappings, pass through as-is
        if not field_mappings:
            mapped = dict(src_row)
        mapped_rows.append(mapped)

    # POST to ontology-service ingest endpoint
    try:
        async with httpx.AsyncClient(timeout=60) as client:
            resp = await client.post(
                f"{ONTOLOGY_API}/object-types/{object_type_id}/records/ingest",
                json={"records": mapped_rows, "pk_field": pk_field},
                headers={"x-tenant-id": tenant_id},
            )
            if not resp.is_success:
                raise HTTPException(
                    status_code=502,
                    detail=f"Ontology ingest failed: {resp.status_code} - {resp.text[:500]}",
                )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Failed to reach ontology service: {str(e)}")

    # Clear staging after successful import
    _file_staging.pop(connector_id, None)

    # Update connector sync metadata
    row.last_sync = datetime.now(timezone.utc)
    row.last_sync_row_count = len(mapped_rows)
    await db.commit()

    return {"imported": len(mapped_rows), "message": f"Successfully imported {len(mapped_rows)} records"}


# ── Database Connector Endpoints ──────────────────────────────────────────────


def _get_db_config(row: ConnectorRow) -> dict:
    """Build database connection config from a connector row."""
    credentials = decrypt_credentials(row.credentials)
    return _build_db_config(credentials, row.config)


@router.get("/{connector_id}/tables")
async def list_db_tables(
    connector_id: str,
    x_tenant_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    """List all user tables from a database connector."""
    tenant_id = x_tenant_id or "tenant-001"
    result = await db.execute(
        select(ConnectorRow).where(
            ConnectorRow.id == connector_id,
            ConnectorRow.tenant_id == tenant_id,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Connector not found")
    if row.type not in DB_CONNECTOR_TYPES:
        raise HTTPException(status_code=400, detail=f"Connector type '{row.type}' is not a database connector")

    try:
        config = _get_db_config(row)
        tables = await db_list_tables(row.type, config)
        return {"tables": tables}
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Database error: {str(e)}")


@router.get("/{connector_id}/tables/{table_name}/schema")
async def get_table_schema(
    connector_id: str,
    table_name: str,
    x_tenant_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    """Return column info for a database table."""
    tenant_id = x_tenant_id or "tenant-001"
    result = await db.execute(
        select(ConnectorRow).where(
            ConnectorRow.id == connector_id,
            ConnectorRow.tenant_id == tenant_id,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Connector not found")
    if row.type not in DB_CONNECTOR_TYPES:
        raise HTTPException(status_code=400, detail=f"Connector type '{row.type}' is not a database connector")

    try:
        config = _get_db_config(row)
        columns = await db_table_schema(row.type, config, table_name)
        return {"columns": columns}
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Database error: {str(e)}")


@router.get("/{connector_id}/tables/{table_name}/preview")
async def preview_db_table(
    connector_id: str,
    table_name: str,
    limit: int = Query(100, ge=1, le=10000),
    x_tenant_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    """Preview rows from a database table."""
    tenant_id = x_tenant_id or "tenant-001"
    result = await db.execute(
        select(ConnectorRow).where(
            ConnectorRow.id == connector_id,
            ConnectorRow.tenant_id == tenant_id,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Connector not found")
    if row.type not in DB_CONNECTOR_TYPES:
        raise HTTPException(status_code=400, detail=f"Connector type '{row.type}' is not a database connector")

    try:
        config = _get_db_config(row)
        preview = await db_preview_table(row.type, config, table_name, limit)
        return preview
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Database error: {str(e)}")


@router.get("/{connector_id}/query")
async def query_db(
    connector_id: str,
    query: str = Query(..., description="SQL query to execute"),
    x_tenant_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    """
    Execute a SQL query against a database connector and return the result rows.
    Used by the pipeline SOURCE node for database-type connectors.
    """
    tenant_id = x_tenant_id or "tenant-001"
    result = await db.execute(
        select(ConnectorRow).where(
            ConnectorRow.id == connector_id,
            ConnectorRow.tenant_id == tenant_id,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Connector not found")
    if row.type not in DB_CONNECTOR_TYPES:
        raise HTTPException(status_code=400, detail=f"Connector type '{row.type}' is not a database connector")

    try:
        config = _get_db_config(row)
        rows = await db_run_query(row.type, config, query)
        # Serialize any non-JSON-serializable values (dates, decimals, etc.)
        serialized = []
        for r in rows:
            clean = {}
            for k, v in r.items():
                if isinstance(v, (datetime,)):
                    clean[k] = v.isoformat()
                elif hasattr(v, '__float__'):
                    clean[k] = float(v)
                else:
                    clean[k] = v
            serialized.append(clean)
        return {"rows": serialized, "row_count": len(serialized)}
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Database query error: {str(e)}")


# ── WhatsApp Proxy Routes ────────────────────────────────────────────────────

async def _wa_proxy(method: str, path: str, tenant_id: str, body: dict | None = None, timeout: float = 30):
    """Forward a request to whatsapp-service and return the JSON response."""
    async with httpx.AsyncClient(timeout=timeout) as client:
        kwargs: dict = {"headers": {"x-tenant-id": tenant_id}}
        if body is not None:
            kwargs["json"] = body
        r = await getattr(client, method)(f"{WHATSAPP_API}{path}", **kwargs)
        if not r.is_success:
            raise HTTPException(status_code=r.status_code, detail=r.text[:500])
        return r.json()


@router.post("/{connector_id}/whatsapp/start")
async def wa_start_session(
    connector_id: str,
    x_tenant_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    row = (await db.execute(select(ConnectorRow).where(ConnectorRow.id == connector_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Connector not found")
    tenant = x_tenant_id or row.tenant_id or "tenant-001"
    return await _wa_proxy("post", f"/api/v1/sessions/{connector_id}/start", tenant, {"tenantId": tenant})


@router.delete("/{connector_id}/whatsapp/stop")
async def wa_stop_session(
    connector_id: str,
    x_tenant_id: Optional[str] = Header(None),
):
    return await _wa_proxy("delete", f"/api/v1/sessions/{connector_id}/stop", x_tenant_id or "tenant-001")


@router.post("/{connector_id}/whatsapp/unlink")
async def wa_unlink(
    connector_id: str,
    x_tenant_id: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_session),
):
    """Wipe stored auth and restart — produces a fresh QR. Use when the
    user wants to re-link the device (different phone, stale creds, etc.)."""
    row = (await db.execute(select(ConnectorRow).where(ConnectorRow.id == connector_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Connector not found")
    tenant = x_tenant_id or row.tenant_id or "tenant-001"
    return await _wa_proxy("post", f"/api/v1/sessions/{connector_id}/unlink", tenant, {"tenantId": tenant})


@router.get("/{connector_id}/whatsapp/status")
async def wa_status(
    connector_id: str,
    x_tenant_id: Optional[str] = Header(None),
):
    return await _wa_proxy("get", f"/api/v1/sessions/{connector_id}/status", x_tenant_id or "tenant-001")


@router.get("/{connector_id}/whatsapp/qr")
async def wa_qr_sse(
    connector_id: str,
    x_tenant_id: Optional[str] = Header(None),
):
    """Proxy SSE stream from whatsapp-service for QR codes."""
    from fastapi.responses import StreamingResponse

    async def stream():
        async with httpx.AsyncClient(timeout=120) as client:
            async with client.stream(
                "GET",
                f"{WHATSAPP_API}/api/v1/sessions/{connector_id}/qr",
                headers={"x-tenant-id": x_tenant_id or "tenant-001"},
            ) as r:
                async for line in r.aiter_lines():
                    yield line + "\n"

    return StreamingResponse(stream(), media_type="text/event-stream")


@router.get("/{connector_id}/whatsapp/chats")
async def wa_list_chats(
    connector_id: str,
    type: Optional[str] = Query(None),
    x_tenant_id: Optional[str] = Header(None),
):
    path = f"/api/v1/sessions/{connector_id}/chats"
    if type:
        path += f"?type={type}"
    return await _wa_proxy("get", path, x_tenant_id or "tenant-001")


@router.patch("/{connector_id}/whatsapp/chats/{jid}/monitor")
async def wa_toggle_monitor(
    connector_id: str,
    jid: str,
    x_tenant_id: Optional[str] = Header(None),
    monitored: bool = True,
):
    return await _wa_proxy(
        "patch",
        f"/api/v1/sessions/{connector_id}/chats/{jid}/monitor",
        x_tenant_id or "tenant-001",
        {"monitored": monitored},
    )


@router.post("/{connector_id}/whatsapp/chats/monitor-all")
async def wa_monitor_all(
    connector_id: str,
    x_tenant_id: Optional[str] = Header(None),
    monitored: bool = True,
):
    return await _wa_proxy(
        "post",
        f"/api/v1/sessions/{connector_id}/chats/monitor-all",
        x_tenant_id or "tenant-001",
        {"monitored": monitored},
    )


@router.get("/{connector_id}/whatsapp/messages")
async def wa_messages(
    connector_id: str,
    chat_jid: Optional[str] = Query(None),
    since: Optional[str] = Query(None),
    limit: int = Query(500),
    offset: int = Query(0),
    x_tenant_id: Optional[str] = Header(None),
):
    params = f"?limit={limit}&offset={offset}"
    if chat_jid:
        params += f"&chat_jid={chat_jid}"
    if since:
        params += f"&since={since}"
    return await _wa_proxy("get", f"/api/v1/sessions/{connector_id}/messages{params}", x_tenant_id or "tenant-001")
