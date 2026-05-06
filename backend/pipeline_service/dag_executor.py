"""
DAG Executor — walks the pipeline DAG topology and executes each node on real records.

Connectors are the raw sources. Pipelines transform the data. Object types are the output.

SOURCE     → fetches real rows from the connector service
ENRICH     → per-row lookup against a second connector (fan-out detail calls)
MAP        → field renaming and transforms on actual records
FILTER     → drops rows that don't match a condition
DEDUPE     → deduplicates by primary key field
CAST       → type coercion on field values
FLATTEN    → explodes an array field into one row per item
VALIDATE   → drops rows missing required fields
SINK_OBJECT → pushes transformed records directly to the ontology ingest endpoint
SINK_EVENT  → converts records into process mining events and writes to event-log
AGENT_RUN  → fires a configured AI agent with the batch of records as context; agent
              can call action_propose which lands proposals in the Human Actions queue
"""
import asyncio
import logging
import os
import re
import httpx
from datetime import datetime, timezone
from uuid import uuid4
from typing import Any

logger = logging.getLogger("dag_executor")
from shared.models import Pipeline
from shared.enums import PipelineStatus, NodeType
from shared.token_tracker import track_token_usage

ONTOLOGY_API = os.environ.get("ONTOLOGY_SERVICE_URL", "http://ontology-service:8004")
EVENT_LOG_API = os.environ.get("EVENT_LOG_SERVICE_URL", "http://event-log-service:8005")
CONNECTOR_API = os.environ.get("CONNECTOR_SERVICE_URL", "http://connector-service:8001")
AGENT_API = os.environ.get("AGENT_SERVICE_URL", "http://agent-service:8013")


def _resolve_date_templates(params: dict, last_sync=None) -> dict:
    """Inline resolution of date templates in query params for pipeline executor."""
    from datetime import timedelta
    now = datetime.now(timezone.utc)

    def fmt(dt, f):
        return (f.replace('YYYY', dt.strftime('%Y')).replace('MM', dt.strftime('%m'))
                 .replace('DD', dt.strftime('%d')).replace('HH', dt.strftime('%H'))
                 .replace('mm', dt.strftime('%M')).replace('ss', dt.strftime('%S')))

    result = {}
    for k, v in params.items():
        s = str(v)
        m = re.match(r'^\{\{\$today:(.+)\}\}$', s)
        if m:
            result[k] = fmt(now, m.group(1)); continue
        m = re.match(r'^\{\{\$daysAgo:(\d+):(.+)\}\}$', s)
        if m:
            result[k] = fmt(now - timedelta(days=int(m.group(1))), m.group(2)); continue
        m = re.match(r'^\{\{\$lastRun:(.+)\}\}$', s)
        if m:
            dt = last_sync if last_sync else (now - timedelta(days=7))
            result[k] = fmt(dt, m.group(1)); continue
        result[k] = s
    return result


def _resolve_path_templates(path: str, last_sync=None) -> str:
    """Resolve {{$lastRun}}, {{$lastRun:FORMAT}}, {{$today:FORMAT}}, {{$daysAgo:N:FORMAT}}
    placeholders anywhere in a URL path or query string. Plain {{$lastRun}} defaults to
    ISO 8601 (e.g. 2026-04-23T00:00:00Z), which is what GitHub/Jira/etc expect."""
    from datetime import timedelta
    if not path or "{{" not in path:
        return path
    now = datetime.now(timezone.utc)
    default_fmt = "YYYY-MM-DDTHH:mm:ssZ"

    def fmt(dt, f):
        return (f.replace('YYYY', dt.strftime('%Y')).replace('MM', dt.strftime('%m'))
                 .replace('DD', dt.strftime('%d')).replace('HH', dt.strftime('%H'))
                 .replace('mm', dt.strftime('%M')).replace('ss', dt.strftime('%S'))
                 .replace('Z', 'Z'))

    def sub_today(m):
        return fmt(now, m.group(1))
    def sub_days_ago(m):
        return fmt(now - timedelta(days=int(m.group(1))), m.group(2))
    def sub_last_run(m):
        dt = last_sync if last_sync else (now - timedelta(days=7))
        f = m.group(1) or default_fmt
        return fmt(dt, f)

    path = re.sub(r'\{\{\$today:([^}]+)\}\}', sub_today, path)
    path = re.sub(r'\{\{\$daysAgo:(\d+):([^}]+)\}\}', sub_days_ago, path)
    path = re.sub(r'\{\{\$lastRun(?::([^}]+))?\}\}', sub_last_run, path)
    return path


def _flatten_github_record(record: dict) -> dict:
    """Promote common nested GitHub fields to top-level scalars so Sink nodes
    don't store {login, avatar_url, ...} dicts that render as [object Object]."""
    if not isinstance(record, dict):
        return record
    out = dict(record)

    def _login(obj):
        return obj.get("login") if isinstance(obj, dict) else obj

    # User-like fields → replace dict with login string; keep id/avatar as _id/_avatar
    for field in ("user", "author", "committer", "assignee", "merged_by", "requested_reviewers"):
        v = out.get(field)
        if isinstance(v, dict):
            out[f"{field}_id"] = v.get("id")
            out[f"{field}_avatar_url"] = v.get("avatar_url")
            out[field] = _login(v)
        elif isinstance(v, list):
            out[field] = ", ".join(str(_login(x) or "") for x in v if x) or None

    # Labels: list of {name, color} → comma-joined names
    if isinstance(out.get("labels"), list):
        out["labels"] = ", ".join(l.get("name", "") for l in out["labels"] if isinstance(l, dict))

    # PR head/base: refs + sha
    for side in ("head", "base"):
        v = out.get(side)
        if isinstance(v, dict):
            out[f"{side}_ref"] = v.get("ref")
            out[f"{side}_sha"] = v.get("sha")
            out[f"{side}_repo"] = (v.get("repo") or {}).get("full_name") if isinstance(v.get("repo"), dict) else None
            out.pop(side, None)

    # /repos/.../commits returns a nested commit {message, author:{name,email,date}, ...}
    commit = out.get("commit")
    if isinstance(commit, dict):
        out["commit_message"] = commit.get("message")
        ca = commit.get("author") or {}
        cc = commit.get("committer") or {}
        if isinstance(ca, dict):
            out["commit_author_name"] = ca.get("name")
            out["commit_author_email"] = ca.get("email")
            out["commit_author_date"] = ca.get("date")
        if isinstance(cc, dict):
            out["commit_committer_name"] = cc.get("name")
            out["commit_committer_date"] = cc.get("date")
        out.pop("commit", None)

    # milestone: {title, state, ...} → milestone_title
    m = out.get("milestone")
    if isinstance(m, dict):
        out["milestone_title"] = m.get("title")
        out["milestone"] = m.get("title")

    # Drop deep URL maps that clutter the output
    for junk in ("_links", "pull_request"):
        out.pop(junk, None)

    return out


async def _touch_connector_last_sync(connector_id: str, tenant_id: str):
    """Fire-and-forget: mark the connector's last_sync = now after a successful pipeline run."""
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            await client.patch(
                f"{CONNECTOR_API}/connectors/{connector_id}/last-sync",
                headers={"x-tenant-id": tenant_id},
            )
    except Exception:
        pass


# Max concurrent ENRICH calls per batch to avoid hammering the detail endpoint
_ENRICH_CONCURRENCY = 10


class NodeExecutionResult:
    def __init__(self, node_id: str, rows_in: int, rows_out: int, error: str | None = None):
        self.node_id = node_id
        self.rows_in = rows_in
        self.rows_out = rows_out
        self.error = error


class DagExecutor:
    """Topological DAG walker that executes pipeline nodes on real record data."""

    def _topological_sort(self, pipeline: Pipeline) -> list[str]:
        """Kahn's algorithm for topological sort."""
        adj: dict[str, list[str]] = {n.id: [] for n in pipeline.nodes}
        in_degree: dict[str, int] = {n.id: 0 for n in pipeline.nodes}

        for edge in pipeline.edges:
            adj[edge.source].append(edge.target)
            in_degree[edge.target] = in_degree.get(edge.target, 0) + 1

        queue = [nid for nid, deg in in_degree.items() if deg == 0]
        order = []

        while queue:
            node_id = queue.pop(0)
            order.append(node_id)
            for neighbor in adj.get(node_id, []):
                in_degree[neighbor] -= 1
                if in_degree[neighbor] == 0:
                    queue.append(neighbor)

        return order

    async def execute(
        self,
        pipeline: Pipeline,
        run: dict[str, Any],
        run_list: list[dict],
        progress_callback: Any = None,
    ) -> None:
        """
        Execute the full pipeline DAG. Records flow through each node as real data.
        SINK_OBJECT pushes the final records directly to the ontology service.
        """
        # Structured log line buffer. Each entry: {ts, level, node_id, msg, extra}.
        # Surfaced in the run drilldown view alongside the existing node_audits.
        logs: list[dict] = []

        def _log(level: str, msg: str, node_id: str = "", extra: dict | None = None):
            entry = {
                "ts": datetime.now(timezone.utc).isoformat(),
                "level": level,
                "node_id": node_id,
                "msg": msg,
            }
            if extra:
                entry["extra"] = extra
            logs.append(entry)

        try:
            order = self._topological_sort(pipeline)
            node_map = {n.id: n for n in pipeline.nodes}

            _log("INFO", f"pipeline started · {len(order)} nodes", extra={"trigger": run.get("triggered_by", "api")})

            # When the pipeline has no explicit edges (step-list builder), synthesise
            # sequential edges so records flow from one step to the next in order.
            linear_mode = not pipeline.edges

            # Each node produces a list of real records
            node_records: dict[str, list[dict]] = {}
            # Per-node audit snapshots written into the run record
            node_audits: dict[str, dict] = {}
            source_row_count = 0
            synced_rows = 0

            total_nodes = len(order)
            for idx, node_id in enumerate(order):
                node = node_map.get(node_id)
                if not node:
                    continue

                if linear_mode:
                    # First node is always the source; every subsequent node gets
                    # the output of the previous node.
                    if idx == 0:
                        records_in: list[dict] = []
                        incoming_edges = []
                    else:
                        prev_id = order[idx - 1]
                        records_in = list(node_records.get(prev_id, []))
                        incoming_edges = [prev_id]  # truthy — not an empty list
                else:
                    incoming_edges = [e for e in pipeline.edges if e.target == node_id]
                    if not incoming_edges:
                        records_in = []
                    else:
                        records_in = []
                        for e in incoming_edges:
                            records_in.extend(node_records.get(e.source, []))

                # ── Live progress: surface the current step before we run it.
                run["current_node_id"] = node_id
                run["current_node_label"] = f"{node.type.value} · {node.label}"
                run["current_step_index"] = idx + 1
                run["total_steps"] = total_nodes
                run["node_audits"] = node_audits  # partial — drilldown can already render
                run["logs"] = logs
                if progress_callback:
                    try:
                        await progress_callback(run)
                    except Exception:
                        pass  # progress is best-effort

                t_start = datetime.now(timezone.utc)
                audit_extras: dict = {}
                _log("INFO", f"{node.type.value} starting · {len(records_in)} rows in", node_id=node_id)
                node_error: str | None = None
                # Reset intra-node progress on each node boundary.
                run["current_node_processed"] = None
                run["current_node_total"] = None
                run["current_model"] = None
                run["current_node_meta"] = None
                progress_ctx = {"run": run, "callback": progress_callback} if progress_callback else None
                try:
                    records_out = await _execute_node(
                        node, records_in, pipeline,
                        audit_extras=audit_extras, progress_ctx=progress_ctx,
                    )
                except Exception as node_exc:
                    node_error = str(node_exc)
                    records_out = []
                    audit_extras["error"] = node_error
                    _log("ERROR", f"{node.type.value} raised: {node_error}", node_id=node_id)
                duration_ms = int((datetime.now(timezone.utc) - t_start).total_seconds() * 1000)

                node_records[node_id] = records_out

                dropped_rows = max(0, len(records_in) - len(records_out)) if records_in else 0
                if node_error:
                    pass  # already logged above
                elif dropped_rows > 0:
                    _log("WARN",
                         f"{node.type.value} ok · {len(records_out)} out · {dropped_rows} dropped · {duration_ms}ms",
                         node_id=node_id)
                else:
                    _log("INFO",
                         f"{node.type.value} ok · {len(records_out)} out · {duration_ms}ms",
                         node_id=node_id)
                # Connector / source HTTP errors come through audit_extras
                src_err = audit_extras.get("error") or audit_extras.get("response_error")
                if src_err and not node_error:
                    _log("ERROR", f"{node.type.value} reported error: {src_err}", node_id=node_id,
                         extra={"http_status": audit_extras.get("http_status"), "url": audit_extras.get("url")})

                is_source_node = (linear_mode and idx == 0) or (not linear_mode and not incoming_edges)
                if is_source_node and records_out:
                    source_row_count = len(records_out)

                if node.type in (NodeType.SINK_OBJECT, NodeType.SINK_EVENT):
                    synced_rows = len(records_out)

                # ── Build per-node audit snapshot ──
                dropped = max(0, len(records_in) - len(records_out))

                # Node-specific stats
                stats: dict = {}
                cfg = node.config or {}
                if node.type == NodeType.FILTER:
                    stats = {"expression": cfg.get("expression", ""), "dropped": dropped}
                elif node.type == NodeType.MAP:
                    mappings = cfg.get("mappings")
                    if isinstance(mappings, str):
                        import json as _json, re as _re
                        try:
                            mappings = _json.loads(mappings)
                        except Exception:
                            try:
                                mappings = _json.loads(_re.sub(r',\s*([\}\]])', r'\1', mappings))
                            except Exception:
                                mappings = {}
                    stats = {"mappings": mappings or {}}
                elif node.type == NodeType.ENRICH:
                    matched = len(records_out)
                    total = len(records_in)
                    stats = {
                        "match_rate": round(matched / total, 3) if total else 0,
                        "matched": matched,
                        "unmatched": total - matched,
                        "join_key": cfg.get("joinKey") or cfg.get("join_key", ""),
                    }
                elif node.type == NodeType.DEDUPE:
                    stats = {"duplicates_removed": dropped, "keys": cfg.get("keys", "")}
                elif node.type == NodeType.VALIDATE:
                    stats = {"invalid_dropped": dropped, "required_fields": cfg.get("requiredFields", [])}
                elif node.type == NodeType.CAST:
                    stats = {"casts": cfg.get("casts", {})}
                elif node.type == NodeType.FLATTEN:
                    stats = {"path": cfg.get("path", ""), "expanded_to": len(records_out)}
                elif node.type == NodeType.PIVOT:
                    stats = {"group_by": cfg.get("groupBy") or cfg.get("group_by", ""),
                             "key_field": cfg.get("keyField") or cfg.get("key_field", ""),
                             "value_field": cfg.get("valueField") or cfg.get("value_field", ""),
                             "collapsed_from": len(records_in), "collapsed_to": len(records_out)}
                elif node.type == NodeType.SINK_OBJECT:
                    stats = {"object_type_id": cfg.get("objectTypeId") or node.object_type_id or pipeline.target_object_type_id, "write_mode": cfg.get("write_mode", "upsert")}
                elif node.type == NodeType.SINK_EVENT:
                    stats = {"activity_field": cfg.get("activityField", ""), "case_id_field": cfg.get("caseIdField", "id"), "events_emitted": len(records_out)}
                elif node.type == NodeType.LLM_CLASSIFY:
                    critico = len([r for r in records_out if r.get("llm_prioridad") == "CRITICO"])
                    urgente = len([r for r in records_out if r.get("llm_prioridad") == "URGENTE"])
                    basura = len([r for r in records_out if r.get("llm_categoria") == "BASURA"])
                    operatividad = len([r for r in records_out if r.get("llm_categoria") == "OPERATIVIDAD"])
                    novedad = len([r for r in records_out if r.get("llm_categoria") == "NOVEDAD RELEVANTE"])
                    stats = {
                        "model": cfg.get("model", "claude-haiku-4-5-20251001"),
                        "text_field": cfg.get("textField", "text"),
                        "classified": len(records_out),
                        "critico": critico, "urgente": urgente,
                        "basura": basura, "operatividad": operatividad, "novedad_relevante": novedad,
                    }
                elif node.type == NodeType.ATTACHMENT_PARSE:
                    parsed_total = sum(
                        len(r.get(cfg.get("outputField") or cfg.get("output_field") or "parsed_rows") or [])
                        for r in records_out
                    )
                    stats = {
                        "attachments_field": cfg.get("attachmentsField") or cfg.get("attachments_field", "attachments"),
                        "filename_match": cfg.get("filenameMatch") or cfg.get("filename_match", ""),
                        "output_field": cfg.get("outputField") or cfg.get("output_field", "parsed_rows"),
                        "header_row": int(cfg.get("headerRow") or cfg.get("header_row") or 0),
                        "data_start_col": int(cfg.get("dataStartCol") or cfg.get("data_start_col") or 0),
                        "rows_extracted": parsed_total,
                    }
                elif node.type == NodeType.SOURCE:
                    stats = {
                        "connector_id": cfg.get("connectorId", ""),
                        "configured_endpoint": cfg.get("endpoint", ""),
                        "url": audit_extras.get("url", ""),
                        "http_status": audit_extras.get("http_status"),
                        "resolved_params": audit_extras.get("resolved_params", {}),
                        "raw_row_count": audit_extras.get("raw_row_count"),
                        "response_error": audit_extras.get("response_error"),
                        "error": audit_extras.get("error"),
                        "query": audit_extras.get("query"),
                        "_watermark_value": audit_extras.get("_watermark_value"),
                        # Incremental fetch surface (set by client-side watermark
                        # filter on REST sources; helps confirm the filter is
                        # actually engaging).
                        "incremental_key": audit_extras.get("incremental_key"),
                        "last_watermark": audit_extras.get("last_watermark"),
                        "filtered_row_count": audit_extras.get("filtered_row_count"),
                        "dropped_by_watermark": audit_extras.get("dropped_by_watermark"),
                    }

                # Flatten sample rows: strip large nested arrays to keep payload small
                def _flatten_sample(rows: list[dict], n: int = 5) -> list[dict]:
                    out = []
                    for r in rows[:n]:
                        flat = {}
                        for k, v in r.items():
                            if isinstance(v, list):
                                flat[k] = f"[{len(v)} items]"
                            elif isinstance(v, dict):
                                flat[k] = "{...}"
                            else:
                                flat[k] = v
                        out.append(flat)
                    return out

                node_audits[node_id] = {
                    "node_id": node_id,
                    "node_type": node.type.value,
                    "node_label": node.label,
                    "rows_in": len(records_in),
                    "rows_out": len(records_out),
                    "dropped": dropped,
                    "duration_ms": duration_ms,
                    "started_at": t_start.isoformat(),
                    "sample_in": _flatten_sample(records_in),
                    "sample_out": _flatten_sample(records_out),
                    "stats": stats,
                    "error": node_error,
                }

                # ── Persist incremental progress so the drilldown view shows
                # samples and stats while the pipeline is still running.
                run["node_audits"] = node_audits
                run["logs"] = logs
                run["rows_in"] = source_row_count
                if progress_callback:
                    try:
                        await progress_callback(run)
                    except Exception:
                        pass

            total_out = synced_rows or max((len(r) for r in node_records.values()), default=0)
            finished_at = datetime.now(timezone.utc).isoformat()

            # Propagate watermark values from SOURCE node audit_extras into
            # a top-level key so _get_last_watermark can retrieve it.
            for _na in node_audits.values():
                stats = _na.get("stats") or {}
                wm = stats.get("_watermark_value")
                if wm:
                    node_audits["_watermark_value"] = wm
                    break

            # Check if any node reported an error (e.g. HTTP 4xx/5xx on SOURCE)
            has_node_error = any(
                (na.get("stats") or {}).get("error") or (na.get("stats") or {}).get("response_error")
                for na in node_audits.values()
                if isinstance(na, dict) and "node_id" in na
            )
            final_status = "FAILED" if has_node_error else "COMPLETED"

            _log(
                "ERROR" if final_status == "FAILED" else "OK",
                f"pipeline {final_status.lower()} · {source_row_count} in / {total_out} out",
            )

            run.update({
                "status": final_status,
                "finished_at": finished_at,
                "rows_in": source_row_count,
                "rows_out": total_out,
                "node_audits": node_audits,
                "logs": logs,
            })

            pipeline.status = PipelineStatus.FAILED if has_node_error else PipelineStatus.IDLE
            pipeline.last_run_at = datetime.now(timezone.utc)
            pipeline.last_run_row_count = total_out

            # Emit pipeline event so it appears in the Event Log
            asyncio.create_task(_emit_pipeline_event(
                pipeline_id=pipeline.id,
                pipeline_name=pipeline.name,
                activity=f"PIPELINE_{final_status}",
                timestamp=finished_at,
                rows_in=source_row_count,
                rows_out=total_out,
                status=final_status,
                tenant_id=pipeline.tenant_id or "tenant-001",
            ))

        except Exception as e:
            logger.error("Pipeline %s FAILED: %s", pipeline.id, e, exc_info=True)
            finished_at = datetime.now(timezone.utc).isoformat()
            _log("ERROR", f"pipeline crashed: {e}")
            run.update({
                "status": "FAILED",
                "finished_at": finished_at,
                "error": str(e),
                "logs": logs,
            })
            pipeline.status = PipelineStatus.FAILED

            asyncio.create_task(_emit_pipeline_event(
                pipeline_id=pipeline.id,
                pipeline_name=pipeline.name,
                activity="PIPELINE_FAILED",
                timestamp=finished_at,
                rows_in=0,
                rows_out=0,
                status="FAILED",
                error=str(e),
                tenant_id=pipeline.tenant_id or "tenant-001",
            ))


# ── Node Handlers ─────────────────────────────────────────────────────────────

async def _execute_node(
    node, records_in: list[dict], pipeline: Pipeline,
    audit_extras: dict | None = None,
    progress_ctx: dict | None = None,
) -> list[dict]:
    if node.type == NodeType.SOURCE:
        return await _source(node, pipeline, audit_extras=audit_extras)
    if node.type == NodeType.ENRICH:
        return await _enrich(node, records_in)
    if node.type == NodeType.MAP:
        return _map(node, records_in)
    if node.type == NodeType.FILTER:
        return _filter(node, records_in)
    if node.type == NodeType.DEDUPE:
        return _dedupe(node, records_in)
    if node.type == NodeType.CAST:
        return _cast(node, records_in)
    if node.type == NodeType.FLATTEN:
        return _flatten(node, records_in)
    if node.type == NodeType.PIVOT:
        return _pivot(node, records_in)
    if node.type == NodeType.VALIDATE:
        return _validate(node, records_in)
    if node.type == NodeType.SINK_OBJECT:
        return await _sink_object(node, records_in, pipeline)
    if node.type == NodeType.SINK_EVENT:
        return await _sink_event(node, records_in, pipeline)
    if node.type == NodeType.AGENT_RUN:
        return await _agent_run(node, records_in, pipeline)
    if node.type == NodeType.LLM_CLASSIFY:
        return await _llm_classify(node, records_in, pipeline, progress_ctx=progress_ctx)
    if node.type == NodeType.ATTACHMENT_PARSE:
        return _attachment_parse(node, records_in)
    return records_in


async def _get_last_watermark(pipeline_id: str) -> str | None:
    """Look up the most recent saved watermark for this pipeline, regardless
    of run status. Watermarks are now persisted from the SOURCE node's stats
    on every progress write, so a partial / watchdog-killed run still leaves
    a watermark behind for the next run to pick up from. The presence of a
    non-null watermark_value is the signal that progress was made."""
    try:
        from database import AsyncSessionLocal, PipelineRunRow
        from sqlalchemy import select as sa_select
        async with AsyncSessionLocal() as db:
            result = await db.execute(
                sa_select(PipelineRunRow)
                .where(
                    PipelineRunRow.pipeline_id == pipeline_id,
                    PipelineRunRow.watermark_value.isnot(None),
                )
                .order_by(PipelineRunRow.started_at.desc())
                .limit(1)
            )
            run_row = result.scalar_one_or_none()
            if run_row and run_row.watermark_value:
                return run_row.watermark_value
    except Exception:
        pass
    return None


async def _source(node, pipeline: Pipeline, audit_extras: dict | None = None) -> list[dict]:
    """
    Fetch real records from the configured connector.

    Supports three source modes:
    1. REST API with endpoint — direct HTTP GET with pagination
    2. Database (POSTGRESQL / MYSQL) — SQL query via connector-service /query endpoint
    3. Fallback — connector /schema sample_rows

    When the node config has `incremental: true` and `watermark_column`, the query
    is augmented with a WHERE clause filtering rows newer than the last run's watermark.
    """
    cfg = node.config or {}
    connector_id = (
        cfg.get("connectorId")
        or cfg.get("connector_id")
        or node.connector_id
        or (pipeline.connector_ids[0] if pipeline.connector_ids else None)
    )
    if not connector_id:
        if audit_extras is not None:
            audit_extras["error"] = "No connector_id configured on SOURCE node"
        return []

    endpoint = (cfg.get("endpoint") or "").strip()
    batch_size = int(cfg.get("batchSize") or 500)

    # Always record the configured values so the audit panel shows something even on failure
    if audit_extras is not None:
        audit_extras["connector_id"] = connector_id
        audit_extras["configured_endpoint"] = endpoint

    try:
        async with httpx.AsyncClient(timeout=60) as client:
            # Fetch connector details (base_url + decrypted credentials) via internal endpoint
            conn_r = await client.get(
                f"{CONNECTOR_API}/connectors/{connector_id}/internal",
                headers={
                    "x-tenant-id": pipeline.tenant_id,
                    "x-internal": os.environ.get("INTERNAL_SECRET", "nexus-internal"),
                },
            )
            if not conn_r.is_success:
                if audit_extras is not None:
                    audit_extras["error"] = f"Connector lookup failed: HTTP {conn_r.status_code}"
                return []

            conn = conn_r.json()
            conn_type = conn.get("type", "")
            base_url = (conn.get("base_url") or "").rstrip("/")
            credentials = conn.get("credentials") or {}
            conn_config = conn.get("config") or {}

            # ── GitHub: auto-wire base_url + default headers + auth ───────
            if conn_type == "GITHUB":
                if not base_url:
                    base_url = "https://api.github.com"
                existing_header_keys = {h.get("key", "").lower() for h in conn_config.get("headers", []) if isinstance(h, dict)}
                gh_defaults = [
                    {"key": "Accept", "value": "application/vnd.github+json"},
                    {"key": "X-GitHub-Api-Version", "value": "2022-11-28"},
                ]
                conn_config = {
                    **conn_config,
                    "headers": [
                        *(conn_config.get("headers") or []),
                        *(h for h in gh_defaults if h["key"].lower() not in existing_header_keys),
                    ],
                }
                # GitHub token → bearer
                if credentials.get("token") and not credentials.get("auth_type"):
                    credentials = {**credentials, "auth_type": "bearer_token"}

            # ── Database connector path (POSTGRESQL / MYSQL) ─────────────
            if conn_type in ("POSTGRESQL", "MYSQL"):
                table = cfg.get("table") or conn_config.get("table", "")
                query = cfg.get("query") or conn_config.get("query", "")

                if not query and table:
                    query = f"SELECT * FROM \"{table}\" LIMIT {batch_size}"

                if not query:
                    if audit_extras is not None:
                        audit_extras["error"] = "Database SOURCE requires 'table' or 'query' in node config"
                    return []

                # Incremental watermark support — any of the common config
                # keys being non-empty enables it (the UI only exposes the
                # key, not a separate boolean).
                incremental_key = (
                    cfg.get("incrementalKey") or cfg.get("incremental_key")
                    or cfg.get("watermark_column") or cfg.get("watermarkColumn") or ""
                )
                watermark_column = incremental_key
                incremental = bool(incremental_key) or cfg.get("incremental", False)
                if incremental and watermark_column:
                    last_watermark = await _get_last_watermark(pipeline.id)
                    if last_watermark:
                        # Inject WHERE clause for incremental fetch
                        if "WHERE" in query.upper():
                            query = query + f" AND {watermark_column} > '{last_watermark}'"
                        else:
                            # Insert WHERE before ORDER BY / LIMIT if present
                            import re
                            match = re.search(r'\b(ORDER\s+BY|LIMIT|GROUP\s+BY)\b', query, re.IGNORECASE)
                            if match:
                                insert_pos = match.start()
                                query = query[:insert_pos] + f"WHERE {watermark_column} > '{last_watermark}' " + query[insert_pos:]
                            else:
                                query = query + f" WHERE {watermark_column} > '{last_watermark}'"

                    # Bound each incremental run: oldest-first + capped to batch_size,
                    # so the watermark advances run-over-run instead of every run
                    # re-fetching the full backlog (which trips the watchdog).
                    import re as _re
                    if not _re.search(r'\bORDER\s+BY\b', query, _re.IGNORECASE):
                        query = query + f" ORDER BY {watermark_column} ASC"
                    if not _re.search(r'\bLIMIT\b', query, _re.IGNORECASE):
                        query = query + f" LIMIT {batch_size}"

                if audit_extras is not None:
                    audit_extras["url"] = f"{CONNECTOR_API}/connectors/{connector_id}/query"
                    audit_extras["query"] = query

                # Call the connector service /query endpoint
                q_resp = await client.get(
                    f"{CONNECTOR_API}/connectors/{connector_id}/query",
                    params={"query": query},
                    headers={"x-tenant-id": pipeline.tenant_id},
                    timeout=120,
                )
                if audit_extras is not None:
                    audit_extras["http_status"] = q_resp.status_code
                if not q_resp.is_success:
                    if audit_extras is not None:
                        audit_extras["response_error"] = q_resp.text[:500]
                    return []

                rows = q_resp.json().get("rows", [])
                if audit_extras is not None:
                    audit_extras["raw_row_count"] = len(rows)

                # Track watermark value for incremental
                if incremental and watermark_column and rows:
                    watermark_vals = [r.get(watermark_column) for r in rows if r.get(watermark_column) is not None]
                    if watermark_vals:
                        max_watermark = str(max(watermark_vals))
                        if audit_extras is not None:
                            audit_extras["_watermark_value"] = max_watermark

                if rows:
                    asyncio.create_task(_touch_connector_last_sync(connector_id, pipeline.tenant_id))
                return rows

            # ── WhatsApp connector path ────────────────────────────────
            if conn_type == "WHATSAPP":
                wa_api = os.environ.get("WHATSAPP_SERVICE_URL", "http://whatsapp-service:8025")
                wa_params: dict[str, str | int] = {"limit": batch_size, "offset": 0}

                # Treat any of these config keys being non-empty as the
                # signal to run incrementally — the UI exposes only
                # "Incremental Key", not a separate boolean flag, so
                # requiring a separate `incremental: true` was the reason
                # this was silently doing full pulls forever.
                incremental_key = (
                    cfg.get("incrementalKey") or cfg.get("incremental_key")
                    or cfg.get("watermark_column") or cfg.get("watermarkColumn") or ""
                )
                # Default to "timestamp" (the WhatsApp API's natural watermark
                # column) so old configs that only set incremental: true keep
                # working. Pass an empty string in cfg to disable.
                if not incremental_key and cfg.get("incremental"):
                    incremental_key = "timestamp"
                is_incremental = bool(incremental_key)
                watermark_column = incremental_key

                last_watermark: str | None = None
                if is_incremental:
                    last_watermark = await _get_last_watermark(pipeline.id)
                    if last_watermark:
                        wa_params["since"] = last_watermark
                    # Oldest-first so capping the page count still advances the
                    # watermark monotonically. Without this, a backlog of N pages
                    # only ever sees the newest page and old messages are skipped.
                    wa_params["order"] = "asc"

                # Cap incremental runs so a backlog drains over multiple runs
                # instead of one run trying to chew through everything and
                # tripping the 1800s watchdog.
                max_records = int(cfg.get("maxRecords") or cfg.get("max_records") or batch_size)

                wa_url = f"{wa_api}/api/v1/sessions/{connector_id}/messages"
                if audit_extras is not None:
                    audit_extras["url"] = wa_url
                    audit_extras["incremental_key"] = incremental_key or None
                    audit_extras["last_watermark"] = last_watermark

                all_rows: list[dict] = []
                while True:
                    r = await client.get(wa_url, params=wa_params, timeout=60)
                    if not r.is_success:
                        if audit_extras is not None:
                            audit_extras["http_status"] = r.status_code
                            audit_extras["response_error"] = r.text[:500]
                        break
                    data = r.json()
                    page_rows = data.get("rows", [])
                    if not page_rows:
                        break
                    all_rows.extend(page_rows)
                    if is_incremental and len(all_rows) >= max_records:
                        all_rows = all_rows[:max_records]
                        break
                    if len(page_rows) < batch_size:
                        break
                    wa_params["offset"] = int(wa_params.get("offset", 0)) + len(page_rows)

                api_returned = len(all_rows)

                # Belt-and-suspenders: even if the WhatsApp service returned
                # records older than `since` (older builds didn't honor it),
                # filter client-side. With `since` working this is a no-op.
                if is_incremental and last_watermark and all_rows:
                    all_rows = [
                        r for r in all_rows
                        if r.get(watermark_column) is not None
                        and str(r.get(watermark_column)) > str(last_watermark)
                    ]
                dropped_by_watermark = api_returned - len(all_rows)

                # Drop undecryptable / empty messages at source level
                before_text_filter = len(all_rows)
                all_rows = [
                    r for r in all_rows
                    if str(r.get("message_type", "")).lower() != "other"
                    and r.get("text") and str(r["text"]).strip()
                ]

                # Add standard field aliases so downstream nodes work with
                # either raw API names (text, id, timestamp) or semantic
                # names (message_text, message_id, sent_at).
                for row in all_rows:
                    if "text" in row and "message_text" not in row:
                        row["message_text"] = row["text"]
                    if "id" in row and "message_id" not in row:
                        row["message_id"] = row["id"]
                    if "timestamp" in row and "sent_at" not in row:
                        row["sent_at"] = row["timestamp"]

                if audit_extras is not None:
                    audit_extras["http_status"] = 200
                    audit_extras["raw_row_count"] = api_returned
                    audit_extras["filtered_row_count"] = len(all_rows)
                    audit_extras["dropped_no_text"] = before_text_filter - len(all_rows)
                    audit_extras["dropped_by_watermark"] = dropped_by_watermark

                # Track watermark for incremental — uses the kept rows so the
                # next run picks up exactly where this one left off.
                if is_incremental and watermark_column and all_rows:
                    wm_vals = [r.get(watermark_column) for r in all_rows if r.get(watermark_column)]
                    if wm_vals:
                        if audit_extras is not None:
                            audit_extras["_watermark_value"] = str(max(wm_vals))

                if all_rows:
                    asyncio.create_task(_touch_connector_last_sync(connector_id, pipeline.tenant_id))
                return all_rows

            # ── Email (IMAP) connector path ────────────────────────────────
            # Pulls messages via the connector-service /emails/fetch endpoint.
            # Supports incremental on `received_at` and an opt-in attachment
            # payload mode for downstream ATTACHMENT_PARSE.
            if conn_type == "EMAIL_INBOX":
                email_folder = (
                    cfg.get("folder")
                    or conn_config.get("default_folder")
                    or "INBOX"
                )
                include_attach = bool(
                    cfg.get("includeAttachments")
                    or cfg.get("include_attachments")
                    or False
                )
                max_attach = int(
                    cfg.get("maxAttachmentBytes")
                    or cfg.get("max_attachment_bytes")
                    or 10 * 1024 * 1024
                )

                incremental_key = (
                    cfg.get("incrementalKey") or cfg.get("incremental_key")
                    or cfg.get("watermark_column") or cfg.get("watermarkColumn") or ""
                )
                if not incremental_key and cfg.get("incremental"):
                    incremental_key = "received_at"
                is_incremental = bool(incremental_key)
                watermark_column = incremental_key

                last_watermark: str | None = None
                if is_incremental:
                    last_watermark = await _get_last_watermark(pipeline.id)

                params: dict[str, str | int | bool] = {
                    "folder": email_folder,
                    "limit": batch_size,
                    "include_attachments": str(include_attach).lower(),
                    "max_attachment_bytes": max_attach,
                }
                if last_watermark:
                    params["since"] = last_watermark

                fetch_url = f"{CONNECTOR_API}/connectors/{connector_id}/emails/fetch"
                if audit_extras is not None:
                    audit_extras["url"] = fetch_url
                    audit_extras["folder"] = email_folder
                    audit_extras["include_attachments"] = include_attach
                    audit_extras["incremental_key"] = incremental_key or None
                    audit_extras["last_watermark"] = last_watermark

                em_resp = await client.get(
                    fetch_url,
                    params=params,
                    headers={
                        "x-tenant-id": pipeline.tenant_id,
                        "x-internal": os.environ.get("INTERNAL_SECRET", "nexus-internal"),
                    },
                    timeout=180,
                )
                if audit_extras is not None:
                    audit_extras["http_status"] = em_resp.status_code
                if not em_resp.is_success:
                    if audit_extras is not None:
                        audit_extras["response_error"] = em_resp.text[:500]
                    return []

                rows = em_resp.json().get("rows", [])
                api_returned = len(rows)

                # Watermark filtering belt-and-suspenders. IMAP `SINCE` is
                # day-granular, so a same-day row from the previous run can
                # appear again. Filter on the actual ISO timestamp.
                if is_incremental and last_watermark and rows:
                    rows = [
                        r for r in rows
                        if r.get(watermark_column) and str(r.get(watermark_column)) > str(last_watermark)
                    ]

                if is_incremental and watermark_column and rows:
                    wm_vals = [r.get(watermark_column) for r in rows if r.get(watermark_column)]
                    if wm_vals and audit_extras is not None:
                        audit_extras["_watermark_value"] = str(max(wm_vals))

                if audit_extras is not None:
                    audit_extras["raw_row_count"] = api_returned
                    audit_extras["filtered_row_count"] = len(rows)

                if rows:
                    asyncio.create_task(_touch_connector_last_sync(connector_id, pipeline.tenant_id))
                return rows

            # ── Grafana / InfluxDB connector path ──────────────────────────
            # Pulls time-series data through Grafana's /api/ds/query proxy.
            # Each (measurement, field) combination produces one row per
            # (device, time-bucket). Configuration:
            #   measurement      InfluxDB measurement (default "alldevices")
            #   field            field key (e.g. "running", "temp")
            #   devices          comma-separated device list (empty = all)
            #   aggregate_every  Flux aggregateWindow size (default "5m")
            #   aggregate_fn     mean | sum | last | max | min (default mean)
            #   lookback_minutes used when no incremental watermark
            if conn_type == "GRAFANA_INFLUX":
                ds_uid = (
                    conn_config.get("datasource_uid")
                    or credentials.get("datasource_uid")
                    or cfg.get("datasource_uid") or ""
                )
                bucket = (
                    conn_config.get("bucket") or credentials.get("bucket")
                    or cfg.get("bucket") or ""
                )
                measurement = cfg.get("measurement") or conn_config.get("measurement") or "alldevices"
                # Multi-field support: prefer `fields` (CSV) on either the
                # node or the connector; fall back to the legacy single
                # `field` setting. We pivot all of them into one row per
                # (device, time) so records carry the same multi-column
                # shape the device-level ingestion produced.
                fields_raw = (
                    cfg.get("fields") or conn_config.get("fields")
                    or cfg.get("field") or conn_config.get("field")
                    or "running"
                )
                if isinstance(fields_raw, list):
                    fields_list = [str(f).strip() for f in fields_raw if str(f).strip()]
                else:
                    fields_list = [s.strip() for s in str(fields_raw).split(",") if s.strip()]
                if not fields_list:
                    fields_list = ["running"]
                devices_raw = cfg.get("devices") or conn_config.get("devices") or ""
                devices = [d.strip() for d in str(devices_raw).split(",") if d.strip()] or None
                # Empty `aggregate_every` = raw events at native cadence
                # (matches the original device-level ingestion). Set a value
                # like "5m" only when downsampling is desirable.
                aggregate_every = cfg.get("aggregate_every") or conn_config.get("aggregate_every") or ""
                aggregate_fn = (cfg.get("aggregate_fn") or conn_config.get("aggregate_fn") or "mean").lower()
                lookback_minutes = int(cfg.get("lookback_minutes") or 60)
                insecure = bool(conn_config.get("tls_skip_verify") or credentials.get("tls_skip_verify"))

                if not (base_url and ds_uid and bucket):
                    if audit_extras is not None:
                        audit_extras["error"] = (
                            "GRAFANA_INFLUX needs base_url, datasource_uid, and bucket configured "
                            "on the connector"
                        )
                    return []

                # Auth: prefer Grafana API key; fall back to basic auth.
                import base64 as _b64
                token = credentials.get("api_key") or credentials.get("token") or ""
                if token:
                    auth_header = f"Bearer {token}"
                elif credentials.get("username") and credentials.get("password"):
                    auth_header = "Basic " + _b64.b64encode(
                        f"{credentials['username']}:{credentials['password']}".encode()
                    ).decode()
                else:
                    if audit_extras is not None:
                        audit_extras["error"] = "GRAFANA_INFLUX needs api_key (Grafana token) in credentials"
                    return []

                # Incremental — use the last `time` we successfully ingested
                # as the Flux range start. Falls back to lookback_minutes on
                # first run.
                incremental_key = (
                    cfg.get("incrementalKey") or cfg.get("incremental_key")
                    or cfg.get("watermark_column") or cfg.get("watermarkColumn") or ""
                )
                if not incremental_key and cfg.get("incremental"):
                    incremental_key = "time"
                last_watermark: Optional[str] = None
                if incremental_key:
                    last_watermark = await _get_last_watermark(pipeline.id)

                from datetime import datetime as _dt, timedelta as _td, timezone as _tz
                end_dt = _dt.now(_tz.utc).replace(microsecond=0)
                if last_watermark:
                    try:
                        start_dt = _dt.fromisoformat(last_watermark.replace("Z", "+00:00"))
                    except ValueError:
                        start_dt = end_dt - _td(minutes=lookback_minutes)
                else:
                    start_dt = end_dt - _td(minutes=lookback_minutes)

                flux_parts = [
                    f'from(bucket: "{bucket}")',
                    f'  |> range(start: time(v: "{start_dt.isoformat().replace("+00:00", "Z")}"), '
                    f'stop: time(v: "{end_dt.isoformat().replace("+00:00", "Z")}"))',
                    f'  |> filter(fn: (r) => r["_measurement"] == "{measurement}")',
                ]
                fields_arr = "[" + ", ".join(f'"{f}"' for f in fields_list) + "]"
                flux_parts.append(
                    f'  |> filter(fn: (r) => contains(value: r["_field"], set: {fields_arr}))'
                )
                if devices:
                    arr = "[" + ", ".join(f'"{d}"' for d in devices) + "]"
                    flux_parts.append(
                        f'  |> filter(fn: (r) => contains(value: r["device"], set: {arr}))'
                    )
                if aggregate_every:
                    flux_parts.append(
                        f'  |> aggregateWindow(every: {aggregate_every}, fn: {aggregate_fn}, createEmpty: false)'
                    )
                # Pivot folds the multi-field result into one row per (device,
                # _time) — each requested field becomes its own column.
                flux_parts.append(
                    '  |> pivot(rowKey: ["_time"], columnKey: ["_field"], valueColumn: "_value")'
                )
                flux_parts.append('  |> yield(name: "pivoted")')
                flux = "\n".join(flux_parts)

                gf_url = f"{base_url}/api/ds/query"
                if audit_extras is not None:
                    audit_extras["url"] = gf_url
                    audit_extras["datasource_uid"] = ds_uid
                    audit_extras["measurement"] = measurement
                    audit_extras["field"] = field_name
                    audit_extras["devices"] = devices or []
                    audit_extras["aggregate_every"] = aggregate_every
                    audit_extras["aggregate_fn"] = aggregate_fn
                    audit_extras["range_start"] = start_dt.isoformat()
                    audit_extras["range_stop"] = end_dt.isoformat()
                    audit_extras["last_watermark"] = last_watermark

                async with httpx.AsyncClient(timeout=120, verify=not insecure) as gf_client:
                    gf_resp = await gf_client.post(
                        gf_url,
                        headers={"Content-Type": "application/json", "Authorization": auth_header},
                        json={
                            "queries": [{
                                "refId": "A",
                                "datasource": {"uid": ds_uid, "type": "influxdb"},
                                "query": flux,
                                # Override Grafana's 1001-point safety cap;
                                # otherwise raw-mode pulls truncate silently
                                # before reaching the records reshape.
                                "maxDataPoints": 500_000,
                            }],
                        },
                    )
                if audit_extras is not None:
                    audit_extras["http_status"] = gf_resp.status_code
                if gf_resp.status_code >= 400:
                    if audit_extras is not None:
                        audit_extras["response_error"] = gf_resp.text[:500]
                    return []

                payload = gf_resp.json()
                # Reshape pivoted frames → one record per (device, _time)
                # carrying every requested field as its own top-level column.
                records: list[dict] = []
                for ref in payload.get("results", {}).values():
                    for frame in ref.get("frames", []) or []:
                        fields_meta = frame.get("schema", {}).get("fields", []) or []
                        cols = frame.get("data", {}).get("values", []) or []
                        if len(fields_meta) < 2 or len(cols) < 2 or not cols[0]:
                            continue
                        # Find the _time column (always 0 in practice, but
                        # be defensive about column ordering).
                        time_idx = next(
                            (i for i, f in enumerate(fields_meta) if f.get("type") == "time"),
                            0,
                        )
                        value_indices = [i for i in range(len(fields_meta)) if i != time_idx]
                        if not value_indices:
                            continue
                        # Device tag lives on the value-typed columns post-pivot.
                        device_label = ""
                        for vi in value_indices:
                            lbl = fields_meta[vi].get("labels") or {}
                            if lbl.get("device"):
                                device_label = lbl["device"]
                                break
                        n_rows = len(cols[time_idx])
                        for row_idx in range(n_rows):
                            ts_ms = cols[time_idx][row_idx]
                            if ts_ms is None:
                                continue
                            ts_iso = _dt.fromtimestamp(ts_ms / 1000.0, tz=_tz.utc).isoformat()
                            rec: dict = {
                                "id": f"{device_label}:{ts_iso}",
                                "time": ts_iso,
                                "sensor_name": device_label,
                                "device": device_label,
                            }
                            for vi in value_indices:
                                fname = fields_meta[vi].get("name") or f"col_{vi}"
                                val = cols[vi][row_idx] if row_idx < len(cols[vi]) else None
                                if val is None:
                                    continue
                                ftype = (fields_meta[vi].get("type") or "").lower()
                                if ftype == "number":
                                    try:
                                        rec[fname] = float(val)
                                    except (TypeError, ValueError):
                                        rec[fname] = val
                                else:
                                    rec[fname] = val
                            records.append(rec)

                if audit_extras is not None:
                    audit_extras["raw_row_count"] = len(records)

                # Track watermark for next run — uses the latest `time` in the
                # actual returned rows so partial backfills advance reliably.
                if incremental_key and records:
                    times = [r.get(incremental_key) for r in records if r.get(incremental_key)]
                    if times and audit_extras is not None:
                        audit_extras["_watermark_value"] = str(max(times))

                if records:
                    asyncio.create_task(_touch_connector_last_sync(connector_id, pipeline.tenant_id))
                return records

            # Parse last_sync from connector details for template resolution
            raw_last_sync = conn.get("last_sync")
            last_sync_dt = None
            if raw_last_sync:
                try:
                    from datetime import datetime as _dt
                    last_sync_dt = _dt.fromisoformat(raw_last_sync.replace("Z", "+00:00"))
                except Exception:
                    pass

            # Accept REST path if base_url alone is a full URL (host+path+query),
            # even when endpoint is blank. That's the natural case for connectors
            # whose Base URL already points at the resource.
            if base_url:
                # Substitute connector-level placeholders in the endpoint path
                # (e.g. {owner}/{repo}/{org}/{username}) from conn_config.
                resolved_endpoint = endpoint or ""
                for ph_key in ("owner", "repo", "org", "username"):
                    ph_val = conn_config.get(ph_key) or credentials.get(ph_key)
                    if ph_val:
                        resolved_endpoint = resolved_endpoint.replace("{" + ph_key + "}", str(ph_val))

                # Resolve date templates in the endpoint path itself
                # (supports {{$lastRun}}, {{$lastRun:FORMAT}}, {{$today:FORMAT}}, {{$daysAgo:N:FORMAT}})
                resolved_endpoint = _resolve_path_templates(resolved_endpoint, last_sync=last_sync_dt)

                # Build the URL and auth headers from the connector's credentials
                url = f"{base_url}{resolved_endpoint}"
                headers: dict[str, str] = {"Accept": "application/json"}

                # Resolve connector's configured queryParams (supports {{$lastRun}}, {{$today}}, etc.)
                raw_qp = {}
                if isinstance(conn_config.get("queryParams"), dict):
                    raw_qp = conn_config["queryParams"]
                params = _resolve_date_templates(raw_qp, last_sync=last_sync_dt)

                auth_type = credentials.get("auth_type") or credentials.get("type", "none")
                if auth_type == "api_key":
                    key_name = credentials.get("header_name", "X-API-Key")
                    headers[key_name] = credentials.get("api_key", "")
                elif auth_type == "bearer_token":
                    headers["Authorization"] = f"Bearer {credentials.get('token', '')}"
                elif auth_type == "basic":
                    import base64 as _b64
                    encoded = _b64.b64encode(
                        f"{credentials.get('username','')}:{credentials.get('password','')}".encode()
                    ).decode()
                    headers["Authorization"] = f"Basic {encoded}"

                # Support username/password basic auth stored directly in credentials
                if not credentials.get("auth_type") and credentials.get("username") and credentials.get("password"):
                    import base64 as _b64
                    encoded = _b64.b64encode(
                        f"{credentials['username']}:{credentials['password']}".encode()
                    ).decode()
                    headers["Authorization"] = f"Basic {encoded}"

                # Add any custom headers from connector config. The headers
                # field can live at either `conn.headers` (top-level on the
                # connector record) or `conn.config.headers` (nested in
                # config), AND each can be stored as either a dict
                # ({"x-admin-token": "abc"}) or a list of {key, value}
                # objects. Handle all four shapes — older connectors used
                # the dict form, the connector creation flow saves the list
                # form. Without this normalization, every header was
                # silently dropped, leading to "401 Unauthorized" being
                # parsed as a JSON string and a downstream
                # "'str' object has no attribute 'get'" crash.
                _custom_headers_sources = [
                    conn_config.get("headers"),
                    conn.get("headers"),
                ]
                for src in _custom_headers_sources:
                    if isinstance(src, dict):
                        for k, v in src.items():
                            if k:
                                headers[str(k)] = str(v if v is not None else "")
                    elif isinstance(src, list):
                        for extra_h in src:
                            if isinstance(extra_h, dict) and extra_h.get("key"):
                                headers[str(extra_h["key"])] = str(extra_h.get("value", ""))

                # ── Fetch (single call or paginated) ─────────────────────
                # If pagination is disabled or the connector has no pagination
                # strategy, make a single request without injecting limit/offset.
                all_rows: list[dict] = []
                page_limit = batch_size  # rows per page request
                offset = 0
                page_num = 1  # used when pagination_strategy == "page"
                first_call = True
                http_method = (cfg.get("method") or cfg.get("http_method") or "GET").upper()
                paginate = cfg.get("paginate", True)
                # Auto-detect: if connector has no pagination strategy or is set to "none", skip pagination
                pagination_strategy = (
                    conn_config.get("paginationStrategy")
                    or conn_config.get("pagination_strategy")
                    or conn.get("pagination_strategy")
                    or ""
                ).lower()
                if pagination_strategy in ("none", ""):
                    paginate = False
                # GitHub defaults to page-based pagination, capped at 100 per page
                if conn_type == "GITHUB" and paginate:
                    pagination_strategy = "page"
                    page_limit = min(page_limit, 100)

                # Honor the connector's verify_ssl flag for the outbound REST call
                _verify_ssl = conn_config.get("verify_ssl", True)
                if isinstance(_verify_ssl, str):
                    _verify_ssl = _verify_ssl.lower() not in ("false", "0", "no")
                rest_client = httpx.AsyncClient(timeout=60, verify=bool(_verify_ssl))
                while True:
                    page_params = dict(params)
                    if paginate:
                        if pagination_strategy == "page":
                            page_params["page"] = page_num
                            page_params["per_page"] = page_limit
                        else:
                            page_params["limit"] = page_limit
                            page_params["offset"] = offset
                    if http_method == "POST":
                        r = await rest_client.post(url, headers=headers, params=page_params, timeout=60)
                    else:
                        r = await rest_client.get(url, headers=headers, params=page_params, timeout=60)
                    if first_call and audit_extras is not None:
                        audit_extras["url"] = str(r.url)
                        audit_extras["http_status"] = r.status_code
                        audit_extras["resolved_params"] = dict(page_params)
                    first_call = False
                    if not r.is_success:
                        if audit_extras is not None:
                            audit_extras["response_error"] = r.text[:500]
                        break
                    try:
                        data = r.json()
                    except Exception as _je:
                        # Server returned a non-JSON body (HTML error page,
                        # plain text, etc.). Surface the body in the audit
                        # so the user can see why instead of getting a
                        # generic "'str' object has no attribute 'get'".
                        if audit_extras is not None:
                            audit_extras["response_error"] = (
                                f"Response was not valid JSON: {_je}. Body head: {r.text[:300]}"
                            )
                        break
                    if isinstance(data, str):
                        # Some APIs (esp. when auth fails) return a plain
                        # JSON-encoded string. Treat that as an error too.
                        if audit_extras is not None:
                            audit_extras["response_error"] = (
                                f"Response was a JSON string, not an object/array: {data[:300]}"
                            )
                        break
                    page_rows: list[dict] | None = None
                    total_declared: int | None = None
                    has_more: bool | None = None
                    if isinstance(data, list):
                        page_rows = data
                    elif isinstance(data, dict):
                        # dict_unwrap_path: takes a dict-of-arrays at the given path and
                        # flattens it into a single array, injecting each key as a field.
                        # e.g. {sensors: {A: [{t,v}], B: [{t,v}]}} with
                        #      dict_unwrap_path=sensors, group_key_field=sensor_name →
                        #      [{sensor_name: A, t, v}, {sensor_name: B, t, v}]
                        dict_path = cfg.get("dict_unwrap_path") or cfg.get("dictUnwrapPath")
                        if dict_path:
                            group_key_field = cfg.get("group_key_field") or cfg.get("groupKeyField", "group_key")
                            obj = data
                            for part in str(dict_path).split("."):
                                if isinstance(obj, dict):
                                    obj = obj.get(part)
                                else:
                                    obj = None
                                    break
                            if isinstance(obj, dict):
                                flat: list[dict] = []
                                for k, v in obj.items():
                                    if isinstance(v, list):
                                        for row in v:
                                            if isinstance(row, dict):
                                                flat.append({group_key_field: k, **row})
                                    elif isinstance(v, dict):
                                        flat.append({group_key_field: k, **v})
                                page_rows = flat

                        # Check well-known wrapper keys next
                        if page_rows is None:
                            for key in ("data", "results", "items", "records", "value", "rows"):
                                if isinstance(data.get(key), list):
                                    page_rows = data[key]
                                    break
                        # If none matched, check node config for a custom records_path
                        records_path_cfg = cfg.get("records_path") or cfg.get("recordsPath")
                        if page_rows is None and records_path_cfg:
                            rp = records_path_cfg
                            obj = data
                            for part in rp.split("."):
                                if isinstance(obj, dict):
                                    obj = obj.get(part)
                                else:
                                    obj = None
                                    break
                            if isinstance(obj, list):
                                page_rows = obj
                        # Last resort: find the first list value in the response
                        if page_rows is None:
                            for v in data.values():
                                if isinstance(v, list):
                                    page_rows = v
                                    break
                        total_declared = data.get("total")
                        has_more = data.get("has_more")
                    if not page_rows:
                        break
                    all_rows.extend(page_rows)
                    # Hard cap: never hold more than max_rows in memory (default 10k)
                    max_rows = int(cfg.get("max_rows") or cfg.get("maxRows") or 10000)
                    if len(all_rows) >= max_rows:
                        all_rows = all_rows[:max_rows]
                        if audit_extras is not None:
                            audit_extras["truncated_at"] = max_rows
                        break
                    # No pagination — single request only
                    if not paginate:
                        break
                    # Stop if the API signals no more pages
                    if has_more is False:
                        break
                    # Stop if we've reached the declared total
                    if total_declared is not None and len(all_rows) >= total_declared:
                        break
                    # Stop if this page was smaller than requested (last page)
                    if len(page_rows) < page_limit:
                        break
                    if pagination_strategy == "page":
                        page_num += 1
                    else:
                        offset += len(page_rows)

                await rest_client.aclose()

                # Client-side incremental filter. REST APIs that don't accept
                # a since/after query param (i.e. they return the full set on
                # every call) need their watermark applied here, after fetch.
                # We drop rows whose incremental key is <= the last completed
                # run's watermark, then track the new max so the next run
                # picks up only what's actually new. ISO-8601 timestamps and
                # zero-padded numeric ids compare correctly as strings; if
                # your key is something else, set it to one that does.
                incremental_key = (
                    cfg.get("incremental_key") or cfg.get("incrementalKey")
                    or cfg.get("watermark_column") or cfg.get("watermarkColumn") or ""
                )
                if incremental_key and all_rows:
                    raw_count = len(all_rows)
                    last_watermark = await _get_last_watermark(pipeline.id)
                    if last_watermark:
                        all_rows = [
                            r for r in all_rows
                            if r.get(incremental_key) is not None
                            and str(r.get(incremental_key)) > str(last_watermark)
                        ]
                    wm_vals = [r.get(incremental_key) for r in all_rows if r.get(incremental_key) is not None]
                    if audit_extras is not None:
                        audit_extras["incremental_key"] = incremental_key
                        audit_extras["last_watermark"] = last_watermark
                        audit_extras["raw_row_count"] = raw_count
                        audit_extras["filtered_row_count"] = len(all_rows)
                        audit_extras["dropped_by_watermark"] = raw_count - len(all_rows)
                        if wm_vals:
                            audit_extras["_watermark_value"] = str(max(wm_vals))

                if all_rows:
                    if audit_extras is not None and "raw_row_count" not in audit_extras:
                        audit_extras["raw_row_count"] = len(all_rows)
                    if conn_type == "GITHUB":
                        all_rows = [_flatten_github_record(r) for r in all_rows]
                    asyncio.create_task(_touch_connector_last_sync(connector_id, pipeline.tenant_id))
                    return all_rows
                return []

            # No endpoint configured — use /schema which resolves templates via connector service
            schema_url = f"{CONNECTOR_API}/connectors/{connector_id}/schema"
            if audit_extras is not None:
                audit_extras["url"] = schema_url
                audit_extras["http_status"] = None
            r = await client.get(schema_url, headers={"x-tenant-id": pipeline.tenant_id})
            if audit_extras is not None:
                audit_extras["http_status"] = r.status_code
            if r.is_success:
                rows = r.json().get("sample_rows", [])
                if audit_extras is not None:
                    audit_extras["raw_row_count"] = len(rows)
                if rows:
                    asyncio.create_task(_touch_connector_last_sync(connector_id, pipeline.tenant_id))
                return rows
    except Exception as _exc:
        if audit_extras is not None:
            audit_extras["error"] = str(_exc)
    return []


async def _enrich(node, records_in: list[dict]) -> list[dict]:
    """
    Per-row detail lookup against a second connector.

    For each incoming row:
      1. Extract the join key value (e.g. row["id"] = "INC-123")
      2. POST /connectors/{lookupConnectorId}/fetch-row with {"params": {lookupField: "INC-123"}}
      3. Merge the detail response onto the row

    Config fields:
      lookupConnectorId  — the connector that holds the detail endpoint
      joinKey            — field on the incoming row whose value is passed as the lookup param
      lookupField        — query param name on the detail endpoint (defaults to joinKey)
    """
    cfg = node.config or {}
    lookup_connector_id = cfg.get("lookupConnectorId") or cfg.get("lookup_connector_id")
    join_key = cfg.get("joinKey") or cfg.get("join_key", "id")
    # The query param name on the detail endpoint — often same as join_key (e.g. "id")
    lookup_field = cfg.get("lookupField") or cfg.get("lookup_field") or join_key

    if not lookup_connector_id or not records_in:
        return records_in

    enriched: list[dict] = []

    async def _lookup_one(row: dict) -> dict:
        join_val = row.get(join_key)
        if join_val is None:
            return row
        try:
            async with httpx.AsyncClient(timeout=30) as client:
                r = await client.post(
                    f"{CONNECTOR_API}/connectors/{lookup_connector_id}/fetch-row",
                    json={"params": {lookup_field: str(join_val)}},
                    headers={"x-tenant-id": pipeline.tenant_id or "tenant-001"},
                )
                if r.is_success:
                    detail = r.json().get("row", {})
                    return {**row, **detail}
        except Exception:
            pass
        return row

    # Run lookups in concurrent batches to avoid overwhelming the detail endpoint
    for i in range(0, len(records_in), _ENRICH_CONCURRENCY):
        batch = records_in[i:i + _ENRICH_CONCURRENCY]
        results = await asyncio.gather(*[_lookup_one(row) for row in batch])
        enriched.extend(results)

    return enriched


def _get_nested(record: dict, path: str) -> Any:
    """Resolve a dot-notation path against a nested dict (e.g. 'address.city')."""
    parts = path.split(".")
    val: Any = record
    for part in parts:
        if not isinstance(val, dict):
            return None
        val = val.get(part)
    return val


def _map(node, records_in: list[dict]) -> list[dict]:
    """Apply field renaming and transforms to each record.

    Supports two config shapes:
      1. mappings: {"source_field": "target_field", ...}   ← new Pipeline Builder format
         Dotted source paths (e.g. "address.city") are resolved against nested dicts.
         Only the mapped fields are kept in the output record.
      2. transforms: [{source_field, target_field, transform_type}]  ← legacy format
         All existing fields are kept; transforms add/rename specific fields.
    """
    import json as _json

    cfg = node.config or {}

    # ── New format: mappings dict ─────────────────────────────────────────────
    # Value shapes supported:
    #   "source": "target"                           — plain rename
    #   "source": {"target": "t", "transform": "x"} — rename + transform
    #   "source": {"target": "t", "transform": "x", "args": {...}}  — with args
    #   "source": ["t1", {"target": "t2", "transform": "x"}]        — one source → many targets
    mappings_raw = cfg.get("mappings")
    if isinstance(mappings_raw, str):
        try:
            mappings_raw = _json.loads(mappings_raw)
        except Exception:
            # Lenient parse: strip trailing commas before closing braces/brackets
            import re as _re
            cleaned = _re.sub(r',\s*([\}\]])', r'\1', mappings_raw)
            try:
                mappings_raw = _json.loads(cleaned)
            except Exception:
                mappings_raw = None
    if isinstance(mappings_raw, dict) and mappings_raw:
        result = []
        for rec in records_in:
            new_rec: dict = {}
            for src_path, tgt_spec in mappings_raw.items():
                # Normalise to a list of specs so we handle arrays uniformly
                specs = tgt_spec if isinstance(tgt_spec, list) else [tgt_spec]
                for spec in specs:
                    if isinstance(spec, str):
                        val = _get_nested(rec, src_path)
                        if val is not None:
                            new_rec[spec] = val
                    elif isinstance(spec, dict):
                        tgt_field = spec.get("target") or spec.get("targetField")
                        transform = spec.get("transform") or spec.get("transform_type", "")
                        t_args = spec.get("args") or {}
                        if not tgt_field:
                            continue
                        # Build a throwaway record so _apply_transform can act on it
                        tmp = dict(rec)
                        tmp = _apply_transform(tmp, src_path, tgt_field, transform, t_args)
                        if tmp.get(tgt_field) is not None:
                            new_rec[tgt_field] = tmp[tgt_field]
            result.append(new_rec)
        return result

    # ── Legacy format: transforms array ───────────────────────────────────────
    transforms: list[dict] = list(cfg.get("transforms") or [])

    if not transforms:
        sf = cfg.get("sourceField") or cfg.get("source_field", "")
        tf = cfg.get("targetField") or cfg.get("target_field", "")
        tt = cfg.get("transformType") or cfg.get("transform_type", "")
        if sf and tf:
            transforms = [{"source_field": sf, "target_field": tf, "transform_type": tt}]

    if not transforms:
        jke = cfg.get("join_key_extraction")
        if jke:
            transforms = [{
                "source_field": jke.get("source_field", ""),
                "target_field": jke.get("output_field", "__join_key__"),
                "transform_type": jke.get("transform", ""),
            }]

    if not transforms:
        return records_in

    result = []
    for rec in records_in:
        for t in transforms:
            rec = _apply_transform(
                rec,
                t.get("source_field", ""),
                t.get("target_field", ""),
                t.get("transform_type", ""),
            )
        result.append(rec)
    return result


def _filter(node, records_in: list[dict]) -> list[dict]:
    """Drop rows that don't satisfy the configured condition."""
    cfg = node.config or {}
    field = cfg.get("field", "")
    operator = cfg.get("operator", "exists")
    value = cfg.get("value")

    if not field:
        return records_in

    result = []
    for rec in records_in:
        fval = rec.get(field)
        if operator == "exists":
            if fval is not None and fval != "" and fval != []:
                result.append(rec)
        elif operator == "not_null":
            if fval is not None:
                result.append(rec)
        elif operator == "eq":
            if str(fval) == str(value):
                result.append(rec)
        elif operator == "neq":
            if str(fval) != str(value):
                result.append(rec)
        elif operator == "contains":
            # Case-insensitive — email subjects, file names, status strings,
            # etc. arrive with arbitrary casing. A case-sensitive filter is
            # almost always a footgun (e.g. "Ordenes de Compra" vs.
            # "ORDENES DE COMPRA"). Use `eq`/`neq` when exact match matters.
            if value is not None and str(value).lower() in str(fval or "").lower():
                result.append(rec)
        else:
            result.append(rec)
    return result


def _dedupe(node, records_in: list[dict]) -> list[dict]:
    """Deduplicate by one or more key fields.

    Reads `keys` (comma- or newline-separated list — what the UI exposes as
    "Dedupe Keys"), or falls back to legacy `pkField`/`pk_field`, or guesses.
    """
    if not records_in:
        return records_in
    cfg = node.config or {}

    raw_keys = cfg.get("keys") or cfg.get("dedupe_keys") or ""
    if isinstance(raw_keys, list):
        key_fields = [str(k).strip() for k in raw_keys if str(k).strip()]
    else:
        key_fields = [k.strip() for k in str(raw_keys).replace("\n", ",").split(",") if k.strip()]

    if not key_fields:
        legacy = cfg.get("pkField") or cfg.get("pk_field")
        if legacy:
            key_fields = [legacy]
        else:
            key_fields = [_guess_pk(records_in[0])]

    seen: set[tuple] = set()
    result = []
    for rec in records_in:
        key = tuple(str(rec.get(f, "")) for f in key_fields)
        if key not in seen:
            seen.add(key)
            result.append(rec)
    return result


def _cast(node, records_in: list[dict]) -> list[dict]:
    """Coerce field values to specified types.

    `field: "*"` applies the cast to every field in the record. Useful for
    flattening nested objects ahead of a SINK_OBJECT write so they don't
    end up rendered as [object Object] in the UI. When casting a dict or
    list to string we use json.dumps so the result is valid JSON, not the
    Python repr.
    """
    import json as _json
    cfg = node.config or {}
    casts: list[dict] = cfg.get("casts") or []
    if not casts:
        return records_in

    def _coerce(value, to_type: str):
        if value is None:
            return value
        if to_type == "string":
            if isinstance(value, (dict, list)):
                return _json.dumps(value, ensure_ascii=False, default=str)
            if isinstance(value, str):
                return value
            return str(value)
        if to_type == "integer":
            return int(value)
        if to_type == "float":
            return float(value)
        if to_type == "boolean":
            return bool(value)
        return value

    result = []
    for rec in records_in:
        rec = dict(rec)
        for c in casts:
            field = c.get("field", "")
            # accept either `to` or `toType` for the destination type
            to_type = c.get("to") or c.get("toType") or "string"
            try:
                if field == "*":
                    for k in list(rec.keys()):
                        try:
                            rec[k] = _coerce(rec[k], to_type)
                        except (ValueError, TypeError):
                            pass
                elif field in rec:
                    rec[field] = _coerce(rec[field], to_type)
            except (ValueError, TypeError):
                pass
        result.append(rec)
    return result


def _flatten(node, records_in: list[dict]) -> list[dict]:
    """Explode an array field — one row per array item."""
    cfg = node.config or {}
    array_field = cfg.get("arrayField") or cfg.get("array_field", "")
    if not array_field:
        return records_in
    result = []
    for rec in records_in:
        arr = rec.get(array_field, [])
        if isinstance(arr, list) and arr:
            for item in arr:
                base = {k: v for k, v in rec.items() if k != array_field}
                if isinstance(item, dict):
                    base.update(item)
                else:
                    base[array_field] = item
                result.append(base)
        else:
            result.append(rec)
    return result


def _attachment_parse(node, records_in: list[dict]) -> list[dict]:
    """ATTACHMENT_PARSE — extract tabular data from email/file attachments.

    Each input record may carry an attachments list (default field name:
    `attachments`) of `{filename, content_type, size, bytes_b64}`. For every
    attachment whose filename matches `filenameMatch` (regex, default `.*`),
    the bytes are decoded and parsed:
      - .xlsx / .xlsm → openpyxl
      - .csv          → csv.reader (tab/comma autodetected)
      - .tsv          → csv.reader, tab delimiter
    Other extensions are skipped silently.

    The parsed rows are attached to the input record under `outputField`
    (default `parsed_rows`) as a list of dicts. Pipe through FLATTEN with
    `arrayField=parsed_rows` to explode into one record per row.

    Config:
      attachmentsField  source field on the input record. default 'attachments'
      filenameMatch     regex; only matching filenames are parsed. default '.*\\.xlsx?$'
      headerRow         0-indexed sheet/csv row containing column headers. default 0
      dataStartCol      0-indexed first column with data — skips empty leading
                        cols (e.g. cols A-C blank in some report templates). default 0
      outputField       where parsed rows go on the input record. default 'parsed_rows'
      sheet             optional sheet name for .xlsx (default: first non-empty)
      skipBlankRows     drop rows whose primary-key column is empty. default True
      primaryKeyHeader  header name for "primary key" used by skipBlankRows.
                        default = first non-empty header
    """
    import base64
    import csv
    import io
    import re

    cfg = node.config or {}
    attachments_field = cfg.get("attachmentsField") or cfg.get("attachments_field") or "attachments"
    filename_match = cfg.get("filenameMatch") or cfg.get("filename_match") or r".*\.xlsx?$"
    header_row = int(cfg.get("headerRow") or cfg.get("header_row") or 0)
    data_start_col = int(cfg.get("dataStartCol") or cfg.get("data_start_col") or 0)
    output_field = cfg.get("outputField") or cfg.get("output_field") or "parsed_rows"
    target_sheet = cfg.get("sheet") or None
    skip_blank = cfg.get("skipBlankRows", cfg.get("skip_blank_rows", True))
    if isinstance(skip_blank, str):
        skip_blank = skip_blank.lower() in ("true", "1", "yes")
    primary_key_header = cfg.get("primaryKeyHeader") or cfg.get("primary_key_header") or ""

    try:
        fname_re = re.compile(filename_match, re.IGNORECASE)
    except re.error as exc:
        logger.warning("[ATTACHMENT_PARSE] invalid filenameMatch %r: %s", filename_match, exc)
        fname_re = re.compile(r".*\.xlsx?$", re.IGNORECASE)

    # openpyxl is heavy; import lazily so non-Excel pipelines don't pay the cost.
    _openpyxl = None

    def _xlsx_rows(payload: bytes, ext: str) -> list[dict]:
        nonlocal _openpyxl
        if _openpyxl is None:
            import openpyxl  # type: ignore
            _openpyxl = openpyxl
        wb = _openpyxl.load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
        if target_sheet and target_sheet in wb.sheetnames:
            ws = wb[target_sheet]
        else:
            ws = wb.worksheets[0]
        all_rows = list(ws.iter_rows(values_only=True))
        return _rows_to_dicts(all_rows)

    def _csv_rows(payload: bytes, ext: str) -> list[dict]:
        text = payload.decode("utf-8", errors="replace")
        delimiter = "\t" if ext == ".tsv" else None
        if delimiter is None:
            try:
                dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;|\t")
                delimiter = dialect.delimiter
            except csv.Error:
                delimiter = ","
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        all_rows = [tuple(r) for r in reader]
        return _rows_to_dicts(all_rows)

    def _rows_to_dicts(all_rows: list) -> list[dict]:
        if header_row >= len(all_rows):
            return []
        headers_row = all_rows[header_row]
        # Slice off the leading empty columns (cols A-C in some report layouts).
        headers = [
            (str(h).strip() if h is not None else "")
            for h in list(headers_row)[data_start_col:]
        ]
        # Trim trailing empties so an extra blank column doesn't pollute keys.
        while headers and not headers[-1]:
            headers.pop()
        if not headers:
            return []
        # Resolve which header acts as the "non-empty" sentinel for skipBlankRows.
        sentinel_idx = 0
        if primary_key_header:
            try:
                sentinel_idx = headers.index(primary_key_header)
            except ValueError:
                sentinel_idx = 0

        out: list[dict] = []
        for raw in all_rows[header_row + 1:]:
            cells = list(raw)[data_start_col:data_start_col + len(headers)]
            row_dict = {}
            for h, v in zip(headers, cells):
                if not h:
                    continue
                if isinstance(v, datetime):
                    row_dict[h] = v.isoformat()
                else:
                    row_dict[h] = v
            if skip_blank:
                sentinel_val = (
                    row_dict.get(headers[sentinel_idx]) if sentinel_idx < len(headers) else None
                )
                if sentinel_val is None or (isinstance(sentinel_val, str) and not sentinel_val.strip()):
                    continue
            out.append(row_dict)
        return out

    enriched: list[dict] = []
    for rec in records_in:
        atts = rec.get(attachments_field) or []
        if not isinstance(atts, list):
            enriched.append({**rec, output_field: []})
            continue
        parsed: list[dict] = []
        for att in atts:
            if not isinstance(att, dict):
                continue
            if att.get("skipped"):
                continue
            fname = str(att.get("filename") or "")
            if not fname or not fname_re.search(fname):
                continue
            b64 = att.get("bytes_b64") or ""
            if not b64:
                continue
            try:
                payload = base64.b64decode(b64)
            except Exception as exc:
                logger.warning("[ATTACHMENT_PARSE] base64 decode failed for %s: %s", fname, exc)
                continue
            ext = "." + fname.rsplit(".", 1)[-1].lower() if "." in fname else ""
            try:
                if ext in (".xlsx", ".xlsm"):
                    rows = _xlsx_rows(payload, ext)
                elif ext in (".csv", ".tsv"):
                    rows = _csv_rows(payload, ext)
                else:
                    continue
            except Exception as exc:
                logger.warning("[ATTACHMENT_PARSE] parse failed for %s (%s): %s", fname, ext, exc)
                continue
            # Stamp provenance on each parsed row so the SINK can keep the
            # link back to the source email + attachment.
            for r in rows:
                r.setdefault("_attachment_filename", fname)
                r.setdefault("_attachment_content_type", att.get("content_type", ""))
            parsed.extend(rows)

        enriched.append({**rec, output_field: parsed})

    return enriched


def _pivot(node, records_in: list[dict]) -> list[dict]:
    """Long-format → wide-format pivot.

    Collapses N rows that share the same `groupBy` tuple into ONE row, with
    each row's (keyField, valueField) pair becoming a column on the output.

    Example config:
        groupBy:    ["sensor_name", "time"]
        keyField:   "field"
        valueField: "value"

    Input rows:
        {sensor_name: "Rajadora_3", time: T, field: "rpm",       value: 150.22}
        {sensor_name: "Rajadora_3", time: T, field: "temp",      value: 53.3}
        {sensor_name: "Rajadora_3", time: T, field: "wifi_ok",   value: True}
    Output row:
        {sensor_name: "Rajadora_3", time: T, rpm: 150.22, temp: 53.3, wifi_ok: True}

    If two rows in the same group have the same key, the LAST one wins
    (input order). This matches typical "deduplicate by groupBy then pick
    latest" semantics. If you need explicit precedence, run DEDUPE first.
    """
    cfg = node.config or {}
    group_by_raw = cfg.get("groupBy") or cfg.get("group_by") or ""
    if isinstance(group_by_raw, str):
        group_keys = [k.strip() for k in group_by_raw.replace("\n", ",").split(",") if k.strip()]
    elif isinstance(group_by_raw, list):
        group_keys = [str(k).strip() for k in group_by_raw if str(k).strip()]
    else:
        group_keys = []
    key_field = cfg.get("keyField") or cfg.get("key_field", "")
    value_field = cfg.get("valueField") or cfg.get("value_field", "")

    # Misconfigured pivots should fail loud rather than silently passing
    # data through — silent passthrough is what got the user here in the
    # first place ("why is my data still in long format?").
    if not group_keys or not key_field or not value_field:
        return records_in

    # Stable insertion order — keeps output reproducible. Uses dict (3.7+
    # preserves insertion) keyed by the tuple of group_by values.
    grouped: dict[tuple, dict] = {}
    for rec in records_in:
        if not isinstance(rec, dict):
            continue
        gkey = tuple(rec.get(k) for k in group_keys)
        if gkey not in grouped:
            grouped[gkey] = {k: rec.get(k) for k in group_keys}
        col = rec.get(key_field)
        if col is None or col == "":
            continue
        grouped[gkey][str(col)] = rec.get(value_field)
    return list(grouped.values())


def _validate(node, records_in: list[dict]) -> list[dict]:
    """Drop rows that are missing any required field."""
    cfg = node.config or {}
    required_fields: list[str] = cfg.get("requiredFields") or cfg.get("required_fields") or []
    if not required_fields:
        return records_in
    return [rec for rec in records_in if all(rec.get(f) is not None for f in required_fields)]


async def _sink_object(node, records_in: list[dict], pipeline: Pipeline) -> list[dict]:
    """
    Push the pipeline's transformed records directly into the ontology as object records.
    Uses /records/ingest instead of /records/sync — the pipeline owns the data, not the connector.
    """
    cfg = node.config or {}
    ot_id = (cfg.get("objectTypeId") or cfg.get("object_type_id")
             or node.object_type_id or pipeline.target_object_type_id)

    if not ot_id or not records_in:
        return records_in

    # Array append write mode: attach records_in as a nested array on matching target records
    if cfg.get("write_mode") == "array_append":
        array_field = cfg.get("array_field", "meetings")
        merge_key = cfg.get("merge_key", "name")
        join_key = cfg.get("join_key", "__join_key__")

        # If records are flowing through the pipeline, use them (the correct path)
        if records_in:
            try:
                async with httpx.AsyncClient(timeout=60) as client:
                    resp = await client.post(
                        f"{ONTOLOGY_API}/object-types/{ot_id}/records/array-append",
                        json={
                            "array_field": array_field,
                            "merge_key": merge_key,
                            "join_key": join_key,
                            "records": records_in,
                        },
                        headers={"x-tenant-id": pipeline.tenant_id},
                    )
            except Exception:
                pass
        else:
            # Fallback: re-fetch from connector schema (legacy path)
            await _execute_array_append(
                pipeline=pipeline,
                ot_id=ot_id,
                array_field=array_field,
                merge_key=merge_key,
            )
        return records_in

    # ── Apply pre-write filter conditions ────────────────────────────────────
    raw_conditions = cfg.get("filterConditions")
    if isinstance(raw_conditions, str) and raw_conditions.strip():
        import json as _json
        try:
            raw_conditions = _json.loads(raw_conditions)
        except Exception:
            raw_conditions = []
    if isinstance(raw_conditions, list) and raw_conditions:
        records_in = [r for r in records_in if _apply_conditions(r, raw_conditions)]
    if not records_in:
        return []

    pk_field_raw = (
        cfg.get("mergeKey") or cfg.get("merge_key")
        or cfg.get("pkField") or cfg.get("pk_field")
        or (_guess_pk(records_in[0]) if records_in else "id")
    )

    # Composite PK support: if merge key is comma-separated, synthesize a
    # _composite_pk field on each record by joining values, and use that as
    # the upsert key. e.g. "sensor_name,time,field" → "Afelpadora|2026-..|heap"
    if isinstance(pk_field_raw, str) and "," in pk_field_raw:
        pk_parts = [p.strip() for p in pk_field_raw.split(",") if p.strip()]
        if pk_parts:
            for rec in records_in:
                rec["_composite_pk"] = "|".join(str(rec.get(p, "")) for p in pk_parts)
            pk_field = "_composite_pk"
        else:
            pk_field = pk_field_raw
    else:
        pk_field = pk_field_raw

    # ── Skip diff-based events if a SINK_EVENT node already handles process events ──
    has_sink_event_node = pipeline.nodes and any(
        (getattr(_n, "type", None) or (_n.get("type") if isinstance(_n, dict) else None)) == "SINK_EVENT"
        for _n in pipeline.nodes
    )
    if has_sink_event_node:
        # SINK_EVENT emits stage-transition events — no per-field diffs needed here,
        # so skip the existing-records pre-fetch entirely.
        ingest_concurrency = max(1, int(os.environ.get("SINK_OBJECT_CONCURRENCY", "4")))
        batch_sz = 500
        batches = [records_in[i:i + batch_sz] for i in range(0, len(records_in), batch_sz)]
        sem = asyncio.Semaphore(ingest_concurrency)

        async def _ingest(batch: list[dict]) -> None:
            async with sem:
                try:
                    async with httpx.AsyncClient(timeout=120) as client:
                        await client.post(
                            f"{ONTOLOGY_API}/object-types/{ot_id}/records/ingest",
                            json={"records": batch, "pk_field": pk_field, "pipeline_id": pipeline.id},
                            headers={"x-tenant-id": pipeline.tenant_id},
                        )
                except Exception:
                    pass

        if batches:
            await asyncio.gather(*(_ingest(b) for b in batches))
        return records_in

    # ── Fetch existing records to diff (Celonis-style record-level events) ──
    # Only fetch existing records if we have a small enough batch to diff safely.
    # For large ingestions (>2000 records), skip diffing to avoid OOM.
    existing_by_pk: dict[str, dict] = {}
    skip_diff = len(records_in) > 2000
    if not skip_diff:
        try:
            async with httpx.AsyncClient(timeout=30) as client:
                r = await client.get(
                    f"{ONTOLOGY_API}/object-types/{ot_id}/records",
                    params={"limit": 5000},
                    headers={"x-tenant-id": pipeline.tenant_id},
                )
                if r.is_success:
                    for rec in r.json().get("records", []):
                        key = str(rec.get(pk_field, ""))
                        if key:
                            existing_by_pk[key] = rec
        except Exception:
            pass

    # ── Build per-record events with field-level diffs ──
    connector_id = pipeline.connector_ids[0] if pipeline.connector_ids else ""
    record_events: list[dict] = []
    fallback_now = datetime.now(timezone.utc).isoformat()

    # Fields that carry the record's own business timestamps (tried in order)
    _RECORD_TS_FIELDS = [
        "occurred_at", "timestamp",
        "hs_lastmodifieddate", "lastmodifieddate", "updatedAt", "updated_at",
        "createdate", "createdAt", "created_at", "date",
    ]
    _RECORD_CREATED_TS_FIELDS = [
        "occurred_at", "timestamp",
        "createdate", "createdAt", "created_at", "hs_createdate",
        "hs_lastmodifieddate", "lastmodifieddate",
    ]

    # Which field's value should become the activity name (e.g. "dealstage")
    # Check: (1) sink node config, (2) any SINK_EVENT node on the same pipeline,
    #         (3) auto-detect common stage/status field names from the first record
    _AUTO_ACTIVITY_FIELDS = [
        "activity",  # event-log style (e.g. ClinicalEvent)
        "stage", "status", "dealstage", "deal_stage", "pipeline_stage",
        "hs_dealstage", "phase", "state", "step", "substatus",
    ]
    activity_field = cfg.get("activityField") or cfg.get("activity_field", "")
    if not activity_field and pipeline.nodes:
        for _n in pipeline.nodes:
            _ncfg = (_n.config or {}) if not isinstance(_n, dict) else (_n.get("config") or {})
            _af = _ncfg.get("activityField") or _ncfg.get("activity_field", "")
            if _af:
                activity_field = _af
                break
    if not activity_field and records_in:
        _sample = records_in[0]
        for _af in _AUTO_ACTIVITY_FIELDS:
            if _af in _sample and str(_sample.get(_af, "")).strip():
                activity_field = _af
                break

    def _record_timestamp(rec: dict, prefer_create: bool = False) -> str:
        fields = _RECORD_CREATED_TS_FIELDS if prefer_create else _RECORD_TS_FIELDS
        for f in fields:
            v = rec.get(f)
            if v and str(v).strip() not in ("", "None", "null"):
                return str(v)
        return fallback_now

    for rec in records_in:
        pk_val = str(rec.get(pk_field, ""))
        if not pk_val:
            continue

        existing = existing_by_pk.get(pk_val)
        if existing is None:
            # Brand-new record — use createdate when available
            ts = _record_timestamp(rec, prefer_create=True)
            # If activityField is set and the field has a value, use it as the activity
            if activity_field and rec.get(activity_field):
                raw = str(rec[activity_field])
                # Preserve original casing for event-log activity fields; uppercase stage fields
                activity_name = raw if activity_field == "activity" else raw.upper().replace(" ", "_")
            else:
                activity_name = "RECORD_CREATED"
            record_events.append({
                "id": str(uuid4()),
                "case_id": pk_val,
                "activity": activity_name,
                "timestamp": ts,
                "object_type_id": ot_id,
                "object_id": pk_val,
                "pipeline_id": pipeline.id,
                "connector_id": connector_id,
                "tenant_id": pipeline.tenant_id,
                "attributes": {
                    "pk_field": pk_field,
                    "record_snapshot": {k: v for k, v in rec.items() if not isinstance(v, (list, dict))},
                },
            })
        else:
            # Existing record — compute field-level diffs
            changed_fields = []
            for field, new_val in rec.items():
                if isinstance(new_val, (list, dict)):
                    continue
                old_val = existing.get(field)
                if str(old_val) != str(new_val):
                    changed_fields.append({"field": field, "from": old_val, "to": new_val})

            if not changed_fields:
                continue

            ts = _record_timestamp(rec, prefer_create=False)

            # Check if the activityField changed — if so, emit one event per stage transition
            if activity_field:
                activity_change = next(
                    (cf for cf in changed_fields if cf["field"] == activity_field), None
                )
                if activity_change:
                    # Emit the stage transition as the primary event
                    raw_stage = str(activity_change["to"])
                    new_stage = raw_stage if activity_field == "activity" else raw_stage.upper().replace(" ", "_")
                    record_events.append({
                        "id": str(uuid4()),
                        "case_id": pk_val,
                        "activity": new_stage,
                        "timestamp": ts,
                        "object_type_id": ot_id,
                        "object_id": pk_val,
                        "pipeline_id": pipeline.id,
                        "connector_id": connector_id,
                        "tenant_id": pipeline.tenant_id,
                        "attributes": {
                            "pk_field": pk_field,
                            "from_stage": activity_change["from"],
                            "to_stage": activity_change["to"],
                            "changed_fields": changed_fields,
                        },
                    })
                    # Emit per-field events for everything else that changed (excluding the activityField)
                    for cf in changed_fields:
                        if cf["field"] == activity_field:
                            continue
                        record_events.append({
                            "id": str(uuid4()),
                            "case_id": pk_val,
                            "activity": f"{cf['field'].upper()}_CHANGED",
                            "timestamp": ts,
                            "object_type_id": ot_id,
                            "object_id": pk_val,
                            "pipeline_id": pipeline.id,
                            "connector_id": connector_id,
                            "tenant_id": pipeline.tenant_id,
                            "attributes": {
                                "pk_field": pk_field,
                                "field": cf["field"],
                                "from": cf["from"],
                                "to": cf["to"],
                            },
                        })
                else:
                    # activityField didn't change — emit per-field events
                    for cf in changed_fields:
                        record_events.append({
                            "id": str(uuid4()),
                            "case_id": pk_val,
                            "activity": f"{cf['field'].upper()}_CHANGED",
                            "timestamp": ts,
                            "object_type_id": ot_id,
                            "object_id": pk_val,
                            "pipeline_id": pipeline.id,
                            "connector_id": connector_id,
                            "tenant_id": pipeline.tenant_id,
                            "attributes": {
                                "pk_field": pk_field,
                                "field": cf["field"],
                                "from": cf["from"],
                                "to": cf["to"],
                            },
                        })
            else:
                # No activityField configured — emit one RECORD_UPDATED per record (original behavior)
                record_events.append({
                    "id": str(uuid4()),
                    "case_id": pk_val,
                    "activity": "RECORD_UPDATED",
                    "timestamp": ts,
                    "object_type_id": ot_id,
                    "object_id": pk_val,
                    "pipeline_id": pipeline.id,
                    "connector_id": connector_id,
                    "tenant_id": pipeline.tenant_id,
                    "attributes": {
                        "pk_field": pk_field,
                        "changed_fields": changed_fields,
                        "fields_changed": len(changed_fields),
                    },
                })

    # ── Emit record-level events ──
    if record_events:
        try:
            async with httpx.AsyncClient(timeout=30) as client:
                for i in range(0, len(record_events), 200):
                    await client.post(
                        f"{EVENT_LOG_API}/events/batch",
                        json={"events": record_events[i:i + 200]},
                        headers={"x-tenant-id": pipeline.tenant_id},
                    )
        except Exception:
            pass

    # ── Apply onConflict: skip records that already exist ────────────────────
    on_conflict = cfg.get("onConflict", "overwrite")
    if on_conflict == "skip" and existing_by_pk:
        records_in = [r for r in records_in if str(r.get(pk_field, "")) not in existing_by_pk]
    elif on_conflict == "preserve" and existing_by_pk:
        # Keep existing field values — only write fields not already present
        preserved = []
        for rec in records_in:
            pk_val = str(rec.get(pk_field, ""))
            existing = existing_by_pk.get(pk_val)
            if existing:
                merged = {**rec}
                for k, v in existing.items():
                    if v is not None and v != "":
                        merged[k] = v  # existing value wins
                preserved.append(merged)
            else:
                preserved.append(rec)
        records_in = preserved

    # ── Ingest records into ontology (batched to avoid OOM) ──
    new_source_ids: set[str] = set()
    total_ingested = 0
    total_new = 0
    try:
        async with httpx.AsyncClient(timeout=120) as client:
            batch_sz = 500
            for i in range(0, len(records_in), batch_sz):
                batch = records_in[i:i + batch_sz]
                resp = await client.post(
                    f"{ONTOLOGY_API}/object-types/{ot_id}/records/ingest",
                    json={
                        "records": batch,
                        "pk_field": pk_field,
                        "pipeline_id": pipeline.id,
                    },
                    headers={"x-tenant-id": pipeline.tenant_id},
                )
                if resp.is_success:
                    ingest_data = resp.json()
                    new_source_ids.update(ingest_data.get("new_source_ids", []))
                    total_new += ingest_data.get("new_count", len(batch))
                    total_ingested += ingest_data.get("ingested", len(batch))
            if total_ingested:
                print(f"[SINK_OBJECT] Ingested {total_ingested} records ({total_new} new) in {(len(records_in) + batch_sz - 1) // batch_sz} batches")
    except Exception:
        pass

    # ── Update object type schema + link source ───────────────────────────────
    if records_in:
        try:
            await _update_object_type_schema(
                ot_id=ot_id,
                sample=records_in[0],
                pipeline=pipeline,
                connector_id=connector_id or "",
            )
        except Exception:
            pass

    # Only pass NEW records downstream (to AGENT_RUN etc.) — not already-existing ones
    if new_source_ids:
        new_records = [r for r in records_in if str(r.get(pk_field) or "") in new_source_ids]
        return new_records if new_records else []
    return []


async def _update_object_type_schema(
    ot_id: str,
    sample: dict,
    pipeline: "Pipeline",
    connector_id: str,
) -> None:
    """
    After a SINK_OBJECT ingest, update the object type's properties (schema)
    from the first mapped record and register the pipeline + connector as sources.
    This makes the object type card show the correct field count and source links.
    """
    import json as _json
    from uuid import uuid4 as _uuid4

    def _infer_semantic(name: str, val: Any) -> str:
        n = name.lower()
        if n.endswith("_id") or n == "id":
            return "IDENTIFIER"
        if "email" in n:
            return "EMAIL"
        if "phone" in n or "tel" in n:
            return "PHONE"
        if "url" in n or "website" in n or "link" in n:
            return "URL"
        if "date" in n or n.endswith("_at") or n.startswith("dt"):
            return "DATETIME"
        if "status" in n or "stage" in n or "state" in n:
            return "STATUS"
        if "name" in n:
            return "PERSON_NAME"
        if "address" in n or "street" in n or "city" in n or "zip" in n:
            return "ADDRESS"
        if isinstance(val, bool):
            return "BOOLEAN"
        if isinstance(val, (int, float)):
            return "QUANTITY"
        return "TEXT"

    def _infer_dtype(val: Any) -> str:
        if isinstance(val, bool):
            return "BOOLEAN"
        if isinstance(val, int):
            return "INTEGER"
        if isinstance(val, float):
            return "FLOAT"
        if isinstance(val, (dict, list)):
            return "JSON"
        return "TEXT"

    async with httpx.AsyncClient(timeout=30) as client:
        # 1. Fetch current object type
        r = await client.get(
            f"{ONTOLOGY_API}/object-types/{ot_id}",
            headers={"x-tenant-id": pipeline.tenant_id},
        )
        if not r.is_success:
            return

        ot_data: dict = r.json()
        existing_prop_names: set[str] = {p["name"] for p in ot_data.get("properties", [])}

        # 2. Infer new properties from sample record (skip nested dicts/lists)
        new_props: list[dict] = []
        for field_name, field_val in sample.items():
            if field_name in existing_prop_names:
                continue
            if isinstance(field_val, (dict, list)):
                continue  # skip complex nested structures
            new_props.append({
                "id": str(_uuid4()),
                "name": field_name,
                "display_name": field_name.replace("_", " ").title(),
                "semantic_type": _infer_semantic(field_name, field_val),
                "data_type": _infer_dtype(field_val),
                "pii_level": "NONE",
                "required": False,
                "source_connector_id": connector_id or None,
                "sample_values": [str(field_val)[:64]] if field_val is not None else [],
            })

        # 3. Merge new properties and update source links
        updated_props = ot_data.get("properties", []) + new_props

        # Add connector to source_connector_ids if not already there
        src_ids: list[str] = list(ot_data.get("source_connector_ids", []))
        if connector_id and connector_id not in src_ids:
            src_ids.append(connector_id)

        ot_data["properties"] = updated_props
        ot_data["source_connector_ids"] = src_ids
        ot_data["source_pipeline_id"] = pipeline.id
        ot_data["version"] = ot_data.get("version", 1) + 1

        # 4. PUT the updated object type back
        await client.put(
            f"{ONTOLOGY_API}/object-types/{ot_id}",
            json=ot_data,
            headers={"x-tenant-id": pipeline.tenant_id},
        )


async def _sink_event(node, records_in: list[dict], pipeline: Pipeline) -> list[dict]:
    """Convert records flowing through the pipeline into process mining events.

    Phase 2 — process-aware fields:
      processId            — links emitted events to a Process definition
      caseKeyField         — record field holding the cross-object business key
                              (e.g. patient_mrn). When set, case_id is set from
                              this value AND attributes.case_key mirrors it, so
                              events from different objects can be joined.
      relatedObjectFields  — list of {objectTypeId, recordField} declaring which
                              other objects this event touches. Stored in
                              attributes.related_objects (consumed in Phase 6).
    """
    cfg = node.config or {}
    object_type_id = (
        cfg.get("objectTypeId")
        or node.object_type_id
        or pipeline.target_object_type_id
        or ""
    )
    case_id_field = cfg.get("caseIdField") or cfg.get("case_id_field", "id")
    activity_field = cfg.get("activityField") or cfg.get("activity_field", "")
    timestamp_field = cfg.get("timestampField") or cfg.get("timestamp_field", "")
    process_id = cfg.get("processId") or cfg.get("process_id") or ""
    case_key_field = cfg.get("caseKeyField") or cfg.get("case_key_field") or ""
    related_object_fields = cfg.get("relatedObjectFields") or cfg.get("related_object_fields") or []
    connector_id = pipeline.connector_ids[0] if pipeline.connector_ids else ""

    # If records are flowing through the pipeline, use them directly
    records = records_in if records_in else []

    # Fall back to fetching from ontology if no records came through (e.g. detached SINK_EVENT)
    if not records and object_type_id:
        written = await _emit_events_from_records(
            object_type_id=object_type_id,
            pipeline_id=pipeline.id,
            connector_ids=pipeline.connector_ids,
            case_id_field=case_id_field,
            activity_field=activity_field,
            timestamp_field=timestamp_field,
            tenant_id=pipeline.tenant_id or "tenant-001",
        )
        return [{}] * written

    if not records:
        return []

    def _pick_ts(record: dict, preferred: str) -> str:
        if preferred and record.get(preferred):
            return str(record[preferred])
        for key in ("createdate", "hs_lastmodifieddate", "closedate", "created_at",
                    "updated_at", "timestamp", "date", "occurred_at"):
            if record.get(key):
                return str(record[key])
        for key, val in record.items():
            if val and any(t in key.lower() for t in ("date", "time", "_at", "stamp")):
                return str(val)
        return datetime.now(timezone.utc).isoformat()

    def _pick_val(record: dict, field: str, fallback: str) -> str:
        if field and record.get(field) is not None:
            return str(record[field])
        return fallback

    events = []
    for record in records:
        source_record_id = str(record.get("id", str(uuid4())))
        legacy_case_id = _pick_val(record, case_id_field, source_record_id)
        # When caseKeyField is set, the cross-object case key wins for case_id
        # so events from different object types naturally cluster on it.
        case_key_value = _pick_val(record, case_key_field, "") if case_key_field else ""
        case_id = case_key_value if case_key_value else legacy_case_id
        activity = _pick_val(record, activity_field, "RECORD_SYNCED")
        ts = _pick_ts(record, timestamp_field)

        if ts and ts.isdigit():
            try:
                ts = datetime.fromtimestamp(int(ts) / 1000, tz=timezone.utc).isoformat()
            except Exception:
                pass

        related_objects = []
        for rof in related_object_fields:
            try:
                ot = (rof.get("objectTypeId") if isinstance(rof, dict) else None) or ""
                fld = (rof.get("recordField") if isinstance(rof, dict) else None) or ""
                if not ot or not fld:
                    continue
                val = record.get(fld)
                if val is None or val == "":
                    continue
                related_objects.append({"object_type_id": ot, "object_id": str(val)})
            except Exception:
                continue

        attributes: dict = {
            k: v for k, v in record.items()
            if k not in (case_id_field, activity_field, timestamp_field, case_key_field)
            and not isinstance(v, (list, dict))
        }
        if process_id:
            attributes["process_id"] = process_id
        if case_key_value:
            attributes["case_key"] = case_key_value
            attributes["source_record_id"] = source_record_id
        if related_objects:
            attributes["related_objects"] = related_objects

        events.append({
            "id": str(uuid4()),
            "case_id": case_id,
            "activity": activity,
            "timestamp": ts or datetime.now(timezone.utc).isoformat(),
            "object_type_id": object_type_id,
            "object_id": case_id,
            "pipeline_id": pipeline.id,
            "connector_id": connector_id,
            "tenant_id": pipeline.tenant_id or "tenant-001",
            "attributes": attributes,
        })

    written = 0
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            for i in range(0, len(events), 200):
                chunk = events[i:i + 200]
                resp = await client.post(
                    f"{EVENT_LOG_API}/events/batch",
                    json={"events": chunk},
                    headers={"x-tenant-id": pipeline.tenant_id or "tenant-001"},
                )
                if resp.is_success:
                    written += len(chunk)
                else:
                    logger.error("SINK_EVENT batch POST failed: status=%s body=%s", resp.status_code, resp.text[:500])
    except Exception as exc:
        logger.error("SINK_EVENT error writing events to event-log-service: %s", exc, exc_info=True)

    logger.info("SINK_EVENT wrote %d / %d events for pipeline %s (tenant %s)", written, len(events), pipeline.id, pipeline.tenant_id)
    return records


# ── Agent Run ─────────────────────────────────────────────────────────────────

async def _agent_run(node, records_in: list[dict], pipeline: Pipeline) -> list[dict]:
    """
    Fire a configured AI agent with the batch of records as context.

    The agent receives all records as a JSON block in its prompt and can call
    action_propose to create Human Action proposals (e.g. urgency alerts).
    This node is a pass-through — it returns records_in unchanged so downstream
    nodes (if any) still receive the full record set.

    Config fields:
      agentId    — ID of the agent to run (required)
      prompt     — instruction prepended to the record batch context
      batchSize  — max records per agent call (default 50, to stay within context)
      runAlways  — if true, fire even when there are no incoming records (default false)
    """
    import json as _json
    cfg = node.config or {}
    agent_id = cfg.get("agentId") or cfg.get("agent_id", "")
    prompt = cfg.get("prompt", "Analyze the following records and propose urgent alerts for any that require immediate human attention.")
    batch_size = int(cfg.get("batchSize") or cfg.get("batch_size") or 50)
    run_always = bool(cfg.get("runAlways") or cfg.get("run_always") or False)

    if not agent_id:
        print(f"[AGENT_RUN] Skipped — no agentId configured on node {node.id}")
        return records_in
    if not records_in and not run_always:
        print(f"[AGENT_RUN] Skipped — no records to process")
        return records_in

    # Truncate records to a safe context size
    records_batch = records_in[:batch_size]
    print(f"[AGENT_RUN] Firing agent {agent_id} with {len(records_batch)} records (pipeline {pipeline.id})")

    # Always include the real current UTC time so agents can calculate ages accurately
    from datetime import datetime, timezone as _tz
    now_utc = datetime.now(_tz.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    # Build the message: prompt + serialized record batch (or standalone prompt if runAlways with no records)
    if records_batch:
        record_context = _json.dumps(records_batch, ensure_ascii=False, default=str, indent=2)
        message = (
            f"Current UTC time: {now_utc}\n\n"
            f"{prompt}\n\n"
            f"The following {len(records_batch)} records were just ingested by a pipeline "
            f"(pipeline_id: {pipeline.id}). Analyze each one:\n\n"
            f"```json\n{record_context}\n```"
        )
    else:
        message = f"Current UTC time: {now_utc}\n\n{prompt}"

    async def _fire():
        try:
            async with httpx.AsyncClient(timeout=180) as client:
                resp = await client.post(
                    f"{AGENT_API}/agents/{agent_id}/test",
                    json={"message": message, "dry_run": False,
                          "pipeline_id": pipeline.id,
                          "pipeline_run_id": getattr(pipeline, 'current_run_id', None)},
                    headers={"x-tenant-id": pipeline.tenant_id},
                )
                if resp.is_success:
                    result = resp.json()
                    print(f"[AGENT_RUN] Agent {agent_id} completed — {result.get('iterations', '?')} iterations, final: {str(result.get('final_text',''))[:120]}")
                else:
                    print(f"[AGENT_RUN] Agent {agent_id} returned {resp.status_code}: {resp.text[:300]}")
        except Exception as e:
            print(f"[AGENT_RUN] Failed to call agent {agent_id}: {e}")

    asyncio.create_task(_fire())
    print(f"[AGENT_RUN] Agent task dispatched (non-blocking)")
    return records_in


async def _llm_classify(node, records_in: list[dict], pipeline: Pipeline, progress_ctx: dict | None = None) -> list[dict]:
    """
    LLM_CLASSIFY — sends each record's text through Claude to extract structured
    fields and merges them back onto the record.

    Supports two modes:
      1. Custom prompt + output schema (general purpose)
      2. Built-in PNC police report classification (Salvadoran police reports)

    Config fields:
      textField     — field containing the text to classify (default: "text")
      prompt        — system prompt for the LLM (has a default for PNC classification)
      outputFields  — comma-separated list of fields to extract (default: PNC fields)
      model         — Claude model to use (default: claude-haiku-4-5-20251001)
      batchSize     — records per LLM call (default: 5, max 10)
      createActions — if true, create Human Actions for CRITICO/URGENTE items (default: true)
    """
    import json as _json
    from shared.llm_router import (
        resolve_provider_for_model, make_async_anthropic_client,
        make_openai_compat_client, OPENAI_COMPAT_TYPES, _to_openai_messages,
    )

    cfg = node.config or {}
    text_field = cfg.get("textField") or cfg.get("text_field") or "text"
    model = cfg.get("model") or "claude-haiku-4-5-20251001"
    batch_size = min(int(cfg.get("batchSize") or cfg.get("batch_size") or 5), 10)
    create_actions = cfg.get("createActions", cfg.get("create_actions", True))
    if isinstance(create_actions, str):
        create_actions = create_actions.lower() in ("true", "1", "yes")

    custom_prompt = (cfg.get("prompt") or "").strip()
    tenant_id = pipeline.tenant_id or "tenant-001"

    # PNC mode = the El Salvador police-report defaults (system prompt, user
    # message framing, output keys, pnc_alert action template). Only kicks in
    # for the MJSP tenant — other tenants would get garbage PNC columns and
    # spurious police actions if they happened to leave the prompt blank.
    # PNC_ALLOWED_TENANTS lets ops widen the allowlist without a code change.
    _pnc_default = "tenant-2e382f99"
    pnc_allowed = {
        t.strip() for t in os.environ.get("PNC_ALLOWED_TENANTS", _pnc_default).split(",") if t.strip()
    }
    tenant_is_pnc = tenant_id in pnc_allowed
    # When the user supplies a custom prompt, drop the PNC-specific user-message
    # framing AND the PNC-specific output-key mapping even on the MJSP tenant —
    # otherwise a PO-extraction prompt still gets a "Clasifica mensajes
    # policiales" user message (steering Claude back to police output) and
    # has its returned keys (supplier, items, …) discarded because the
    # extractor only reads categoria/prioridad/etc. Generic mode merges
    # whatever JSON keys the LLM returns, prefixed with `llm_`.
    is_pnc_mode = tenant_is_pnc and not custom_prompt

    if not custom_prompt and not tenant_is_pnc:
        logger.warning(
            "[LLM_CLASSIFY] tenant=%s has no custom prompt and is not in "
            "PNC_ALLOWED_TENANTS — passing %d records through unchanged. "
            "Set a system prompt on the LLM_CLASSIFY step to enable enrichment.",
            tenant_id, len(records_in),
        )
        return records_in

    # Default PNC system prompt (El Salvador police report classification)
    system_prompt = custom_prompt or """Eres un analista de inteligencia policial de El Salvador. Tu tarea es clasificar mensajes de WhatsApp de grupos policiales y extraer información estructurada.

Para cada mensaje, devuelve un objeto JSON con estos campos:

{
  "categoria": "NOVEDAD RELEVANTE" | "OPERATIVIDAD" | "BASURA",
  "tipo_incidente": "<tipo o null si no aplica>",
  "accion_policial": "<acción>",
  "prioridad": "CRITICO" | "URGENTE" | "IMPORTANTE" | "INFORMATIVA",
  "departamento": "<departamento de El Salvador o null>",
  "municipio": "<municipio o null>",
  "lugar": "<dirección específica o null>",
  "fecha_hora": "<fecha/hora del evento en ISO 8601 o null>",
  "involucrados": {
    "responsables": [{"nombre": "<nombre>", "edad": <edad o null>, "rol": "<rol>"}],
    "victimas": [{"nombre": "<nombre>", "edad": <edad o null>, "sexo": "<M/F o null>"}]
  },
  "hecho": "<resumen en máximo 2 oraciones, sin hora/fecha/lugar, usando 'el imputado', 'la víctima', etc.>",
  "incautaciones": {"droga": null, "arma_fuego": null, "arma_blanca": null, "vehiculo": null, "otros": null}
}

Valores válidos para tipo_incidente:
- Fallecidos: HOMICIDIO, HOMICIDIO AGRAVADO, FEMINICIDIO, PERSONA FALLECIDA (ACCIDENTE DE TRÁNSITO), PERSONA FALLECIDA (SUICIDIO), PERSONA FALLECIDA (AHOGADO), PERSONA FALLECIDA (MUERTE NATURAL), PERSONA FALLECIDA (CAUSA POR DETERMINAR)
- Delitos contra la persona: LESIONES, LESIONES GRAVES, VIOLENCIA INTRAFAMILIAR, VIOLACIÓN, AGRESIÓN SEXUAL, AMENAZAS, SECUESTRO, PRIVACIÓN DE LIBERTAD
- Delitos patrimoniales: ROBO, ROBO DE VEHÍCULO, HURTO, ESTAFA
- Armas y drogas: PORTACIÓN ILÍCITA DE ARMA DE FUEGO, PORTACIÓN ILÍCITA DE ARMA BLANCA, POSESIÓN DE DROGA, TRÁFICO DE DROGA
- Orden público: AGRUPACIONES ILÍCITAS, EXTORSIÓN
- Emergencias: INCENDIO, DERRUMBE, INUNDACIÓN
- Rescates: ACCIDENTE DE TRÁNSITO (SIN FALLECIDOS), TRASLADO DE LESIONADO

Valores para accion_policial: DETENCIÓN, DECOMISO, INSPECCIÓN, PATRULLAJE, CHARLA PREVENTIVA, FORMACIÓN, SEGURIDAD EN CENTRO EDUCATIVO, CONTROL VEHICULAR, RESCATE, TRASLADO, VERIFICACIÓN, OPERACIÓN EXTRACCIÓN, OTROS

Prioridad:
- CRITICO: homicidio, feminicidio, secuestro, masacre, persona fallecida (cualquier causa excepto muerte natural)
- URGENTE: detenciones, lesiones graves, robo con violencia, extorsión, armas, drogas, suicidio, persona fallecida (causa por determinar), ahogado, accidente de tránsito con fallecidos
- IMPORTANTE: hurtos, lesiones leves, accidente sin fallecidos
- INFORMATIVA: patrullajes, charlas, control vehicular, fallecidos por muerte natural confirmada

BASURA: SOLO si el mensaje completo es un saludo sin contenido policial (ej: solo "buenos días", "reportándome", "ok", "recibido").
IMPORTANTE: Si un mensaje COMIENZA con un saludo ("Buen día", "Buenos días") pero CONTIENE un reporte policial (homicidio, detención, operatividad, etc.), NO es BASURA — clasifícalo según su contenido real. Lee el mensaje COMPLETO antes de clasificar.
Si es OPERATIVIDAD (patrullaje, charla, seguridad escolar sin incidentes): devuelve categoria="OPERATIVIDAD" y solo los campos relevantes.

IMPORTANTE: Responde SOLO con el array JSON válido. Sin texto antes ni después. Sin bloques ```json```. Sin explicaciones. Solo el JSON puro."""

    provider_cfg = await resolve_provider_for_model(tenant_id, model)
    if not provider_cfg.api_key and provider_cfg.provider_type != "local":
        logger.error("[LLM_CLASSIFY] No API key for provider %s — skipping classification", provider_cfg.provider_type)
        return records_in

    is_openai = provider_cfg.provider_type in OPENAI_COMPAT_TYPES
    llm_timeout_s = float(os.environ.get("LLM_CLASSIFY_TIMEOUT_S", "300"))
    concurrency = max(1, int(os.environ.get("LLM_CLASSIFY_CONCURRENCY", "1")))
    if is_openai:
        client = make_openai_compat_client(provider_cfg, async_client=True)
    else:
        client = make_async_anthropic_client(provider_cfg)
    logger.info("[LLM_CLASSIFY] Using provider=%s type=%s model=%s", provider_cfg.provider_name or "env", provider_cfg.provider_type, provider_cfg.model)

    # Pre-filter: DROP records with no text entirely (reactions, read receipts,
    # decryption failures with message_type "other", etc.).  These are noise —
    # never send them to the LLM and never include them in the output.
    classifiable = []
    dropped = 0
    for record in records_in:
        text_val = record.get(text_field)
        msg_type = str(record.get("message_type", "")).lower()
        if msg_type == "other":
            dropped += 1
            continue
        if text_val and str(text_val).strip() and str(text_val).strip().lower() != "none":
            classifiable.append(record)
        else:
            dropped += 1
    logger.info(
        f"[LLM_CLASSIFY] {len(classifiable)} classifiable of {len(records_in)} total "
        f"({dropped} dropped — no text or message_type=other). "
        f"batch_size={batch_size} concurrency={concurrency}"
    )
    records_in = classifiable

    batches: list[list[dict]] = [records_in[i:i + batch_size] for i in range(0, len(records_in), batch_size)]
    semaphore = asyncio.Semaphore(concurrency)

    # Live progress: surface total + model up front, then tick `processed`
    # after each batch completes. The Hivemind running card reads this so
    # users can watch records crawl through LLM_CLASSIFY in real time.
    total_to_classify = len(records_in)
    processed_count = 0
    batches_done = 0
    total_batches = len(batches)
    cum_input_tokens = 0
    cum_output_tokens = 0

    def _meta_snapshot() -> dict:
        return {
            "batches_done": batches_done,
            "batches_total": total_batches,
            "batch_size": batch_size,
            "concurrency": concurrency,
            "input_tokens": cum_input_tokens,
            "output_tokens": cum_output_tokens,
            "dropped_prefilter": dropped,
            "provider": provider_cfg.provider_type,
        }

    if progress_ctx:
        run = progress_ctx.get("run") or {}
        cb = progress_ctx.get("callback")
        run["current_node_total"] = total_to_classify
        run["current_node_processed"] = 0
        run["current_model"] = provider_cfg.model
        run["current_node_meta"] = _meta_snapshot()
        if cb:
            try:
                await cb(run)
            except Exception:
                pass

    async def _run_batch(batch_idx: int, batch: list[dict]) -> list[dict]:
        nonlocal cum_input_tokens, cum_output_tokens
        parts = []
        for idx, record in enumerate(batch):
            text = str(record.get(text_field, ""))
            sender = record.get("sender_name") or record.get("sender_jid") or "unknown"
            chat = record.get("chat_name") or record.get("chat_jid") or "unknown"
            parts.append(f"--- MENSAJE {idx + 1} ---\nGrupo: {chat}\nDe: {sender}\n\n{text}\n")

        if is_pnc_mode:
            user_msg = (
                f"Clasifica los siguientes {len(batch)} mensajes de WhatsApp policial. "
                f"Devuelve un array JSON con exactamente {len(batch)} objetos, uno por mensaje, en el mismo orden.\n\n"
                + "\n".join(parts)
            )
        else:
            # Neutral framing — the system prompt drives the schema, this just
            # tells Claude to return one JSON object per input message in order.
            user_msg = (
                f"Process the following {len(batch)} message(s). "
                f"Return a JSON array of exactly {len(batch)} object(s), one per message, in the same order. "
                f"Each object must follow the schema described in the system prompt. "
                f"Respond with ONLY the JSON array — no markdown, no commentary.\n\n"
                + "\n".join(parts)
            )

        try:
            async with semaphore:
                if is_openai:
                    oai_kwargs: dict = dict(
                        model=provider_cfg.model,
                        max_tokens=4096,
                        temperature=0.1,
                        messages=_to_openai_messages(system_prompt, [{"role": "user", "content": user_msg}]),
                    )
                    oai_resp = await asyncio.wait_for(
                        client.chat.completions.create(**oai_kwargs),
                        timeout=llm_timeout_s + 5,
                    )
                    raw_text = (oai_resp.choices[0].message.content or "").strip()
                    usage = getattr(oai_resp, "usage", None)
                    in_tok = getattr(usage, "prompt_tokens", 0) if usage else 0
                    out_tok = getattr(usage, "completion_tokens", 0) if usage else 0
                else:
                    resp = await asyncio.wait_for(
                        client.messages.create(
                            model=provider_cfg.model,
                            max_tokens=4096,
                            temperature=0.1,
                            system=system_prompt,
                            messages=[{"role": "user", "content": user_msg}],
                        ),
                        timeout=llm_timeout_s + 5,
                    )
                    raw_text = resp.content[0].text.strip()
                    in_tok = resp.usage.input_tokens
                    out_tok = resp.usage.output_tokens
            cum_input_tokens += int(in_tok or 0)
            cum_output_tokens += int(out_tok or 0)
            track_token_usage(
                pipeline.tenant_id or "unknown", "pipeline_service", provider_cfg.model,
                in_tok, out_tok,
            )
            json_text = raw_text
            if "```" in json_text:
                match = re.search(r"```(?:json)?\s*([\s\S]*?)```", json_text)
                if match:
                    json_text = match.group(1).strip()
            # Some models wrap the JSON in extra prose — find the first [ or {
            if json_text and json_text[0] not in ("[", "{"):
                bracket = json_text.find("[")
                brace = json_text.find("{")
                start = min(p for p in (bracket, brace) if p >= 0) if max(bracket, brace) >= 0 else -1
                if start >= 0:
                    json_text = json_text[start:]

            parsed = _json.loads(json_text)
            if isinstance(parsed, dict) and not isinstance(parsed, list):
                parsed = [parsed]

            def _scalar(v):
                if v is None:
                    return None
                if isinstance(v, list):
                    return ", ".join(str(x) for x in v) if v else None
                if isinstance(v, dict):
                    return _json.dumps(v, ensure_ascii=False)
                return v

            out: list[dict] = []
            for idx, record in enumerate(batch):
                enriched_record = dict(record)
                if idx < len(parsed):
                    classification = parsed[idx]
                    if not isinstance(classification, dict):
                        # Defensive: LLM returned something other than an object
                        # for this slot (string, list, etc.). Skip enrichment.
                        out.append(enriched_record)
                        continue
                    if is_pnc_mode:
                        enriched_record["llm_categoria"] = _scalar(classification.get("categoria"))
                        enriched_record["llm_tipo_incidente"] = _scalar(classification.get("tipo_incidente"))
                        enriched_record["llm_accion_policial"] = _scalar(classification.get("accion_policial"))
                        enriched_record["llm_prioridad"] = _scalar(classification.get("prioridad"))
                        enriched_record["llm_departamento"] = _scalar(classification.get("departamento"))
                        enriched_record["llm_municipio"] = _scalar(classification.get("municipio"))
                        enriched_record["llm_lugar"] = _scalar(classification.get("lugar"))
                        enriched_record["llm_fecha_hora"] = _scalar(classification.get("fecha_hora"))
                        enriched_record["llm_hecho"] = _scalar(classification.get("hecho"))

                        inv = classification.get("involucrados") or {}
                        responsables = inv.get("responsables") or []
                        victimas = inv.get("victimas") or []
                        enriched_record["llm_responsables"] = _json.dumps(responsables, ensure_ascii=False) if responsables else None
                        enriched_record["llm_victimas"] = _json.dumps(victimas, ensure_ascii=False) if victimas else None

                        inc = classification.get("incautaciones") or {}
                        for k, v in inc.items():
                            if v is not None:
                                enriched_record[f"llm_incautacion_{k}"] = v
                    else:
                        # Generic mode — merge every key the LLM returned.
                        # Prefix with `llm_` so the LLM output stays distinguishable
                        # from source fields and can't accidentally clobber them
                        # (e.g. `text`, `id`, `sent_at`).
                        for key, value in classification.items():
                            if not isinstance(key, str) or not key:
                                continue
                            col = key if key.startswith("llm_") else f"llm_{key}"
                            enriched_record[col] = _scalar(value)
                out.append(enriched_record)

            logger.info(f"[LLM_CLASSIFY] Batch {batch_idx + 1}/{len(batches)}: classified {len(batch)} records")
            return out

        except Exception as e:
            logger.error(f"[LLM_CLASSIFY] Batch {batch_idx + 1}/{len(batches)} failed: {repr(e)}", exc_info=True)
            out = []
            for record in batch:
                r = dict(record)
                r["llm_categoria"] = "ERROR"
                r["llm_error"] = repr(e)[:200]
                out.append(r)
            return out

    async def _run_batch_tracked(i: int, b: list[dict]) -> list[dict]:
        nonlocal processed_count, batches_done
        out = await _run_batch(i, b)
        processed_count += len(b)
        batches_done += 1
        if progress_ctx:
            run = progress_ctx.get("run") or {}
            cb = progress_ctx.get("callback")
            run["current_node_processed"] = processed_count
            run["current_node_meta"] = _meta_snapshot()
            if cb:
                try:
                    await cb(run)
                except Exception:
                    pass
        return out

    # The semaphore inside _run_batch is the actual concurrency gate. Running
    # batches via gather lets LLM_CLASSIFY_CONCURRENCY > 1 actually parallelize;
    # with the default of 1, semaphore acquire serializes so behavior matches
    # the prior sequential loop (and gather preserves input order).
    batch_results = await asyncio.gather(*(_run_batch_tracked(i, b) for i, b in enumerate(batches)))
    enriched: list[dict] = []
    for r in batch_results:
        enriched.extend(r)

    # Create Human Actions for CRITICO/URGENTE items. Only in PNC mode —
    # the action template + dedup keys are police-specific and would write
    # garbage into the actions table for a generic prompt.
    if create_actions and is_pnc_mode:
        def _as_str(v) -> str:
            if v is None:
                return ""
            if isinstance(v, (list, tuple)):
                return ", ".join(str(x) for x in v)
            if isinstance(v, dict):
                return _json.dumps(v, ensure_ascii=False)
            return str(v)

        def _is_urgent(r: dict) -> bool:
            p = _as_str(r.get("llm_prioridad")).upper()
            u = _as_str(r.get("llm_urgencia")).upper()
            c = _as_str(r.get("llm_categoria")).lower()
            return (
                p in ("CRITICO", "URGENTE")
                or u in ("CRITICA", "ALTA")
                or c in ("delito_calle", "muerte")
            )
        urgent_records = [r for r in enriched if _is_urgent(r)]
        if urgent_records:
            asyncio.create_task(
                _create_urgent_actions(urgent_records, pipeline)
            )

    return enriched


async def _create_urgent_actions(records: list[dict], pipeline: Pipeline):
    """Create Human Action proposals for CRITICO/URGENTE classified records."""
    import json as _json
    headers = {"x-tenant-id": pipeline.tenant_id, "Content-Type": "application/json"}

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            # Ensure the pnc_alert action definition exists (idempotent)
            check = await client.get(f"{ONTOLOGY_API}/actions/pnc_alert", headers=headers)
            if check.status_code == 404:
                await client.post(
                    f"{ONTOLOGY_API}/actions",
                    json={
                        "name": "pnc_alert",
                        "description": "Alerta policial clasificada por IA desde WhatsApp — requiere revisión humana",
                        "requires_confirmation": True,
                        "enabled": True,
                        "input_schema": {
                            "titulo": "string",
                            "prioridad": "string",
                            "tipo_incidente": "string",
                            "departamento": "string",
                            "hecho": "string",
                            "detalles": "string",
                        },
                    },
                    headers=headers,
                )

            # Fetch existing executions to deduplicate — skip records that already have an action
            existing_source_ids: set[str] = set()
            try:
                ex_resp = await client.get(
                    f"{ONTOLOGY_API}/actions/pnc_alert/executions",
                    params={"limit": 500},
                    headers=headers,
                )
                if ex_resp.status_code == 200:
                    for ex in ex_resp.json():
                        sid = ex.get("source_id")
                        if sid:
                            existing_source_ids.add(sid)
            except Exception as ex_err:
                logger.warning(f"[LLM_CLASSIFY] Could not fetch existing executions for dedup: {ex_err}")

            created = 0
            skipped = 0
            for record in records:
                # Use the record's unique id (message_id) as the dedup key
                record_id = str(record.get("id") or record.get("message_id") or "")
                if record_id and record_id in existing_source_ids:
                    skipped += 1
                    continue

                prioridad = record.get("llm_prioridad", "URGENTE")
                tipo = record.get("llm_tipo_incidente", "Desconocido")
                hecho = record.get("llm_hecho", record.get("text", "")[:200])
                sender = record.get("sender_name") or record.get("sender_jid") or "desconocido"
                chat = record.get("chat_name") or record.get("chat_jid") or "desconocido"
                depto = record.get("llm_departamento") or "N/D"

                title = f"[{prioridad}] {tipo} — {depto}"
                details = (
                    f"Municipio: {record.get('llm_municipio', 'N/D')}\n"
                    f"Lugar: {record.get('llm_lugar', 'N/D')}\n"
                    f"Grupo: {chat}\n"
                    f"Reportó: {sender}\n"
                    f"Responsables: {record.get('llm_responsables', 'N/D')}\n"
                    f"Víctimas: {record.get('llm_victimas', 'N/D')}"
                )

                resp = await client.post(
                    f"{ONTOLOGY_API}/actions/pnc_alert/execute",
                    json={
                        "inputs": {
                            "titulo": title,
                            "prioridad": prioridad,
                            "tipo_incidente": tipo,
                            "departamento": depto,
                            "hecho": hecho,
                            "detalles": details,
                        },
                        "executed_by": f"pipeline:{pipeline.id}",
                        "source": "llm_classify",
                        "source_id": record_id or pipeline.id,
                        "reasoning": f"Clasificación automática de mensaje WhatsApp: {tipo} ({prioridad})",
                    },
                    headers=headers,
                )
                if resp.status_code >= 400:
                    logger.error(f"[LLM_CLASSIFY] Action execute failed ({resp.status_code}): {resp.text[:300]}")
                else:
                    created += 1

            logger.info(f"[LLM_CLASSIFY] Actions: {created} created, {skipped} skipped (already exist)")
    except Exception as e:
        logger.error(f"[LLM_CLASSIFY] Failed to create actions: {e}")


# ── Shared Helpers ────────────────────────────────────────────────────────────

def _guess_pk(record: dict) -> str:
    for candidate in ("hs_object_id", "id", "record_id", "uuid", "incident_id"):
        if record.get(candidate):
            return candidate
    return next(iter(record), "id")


def _apply_conditions(record: dict, conditions: list[dict]) -> bool:
    """
    Return True if all conditions are satisfied by the record.
    Each condition: {field, operator, value}
    Operators: eq, neq, contains, not_contains, gt, lt, gte, lte, is_null, not_null, exists
    """
    for cond in conditions:
        field = cond.get("field", "")
        operator = cond.get("operator", "exists")
        expected = cond.get("value", "")
        actual = record.get(field)

        if operator == "exists":
            if actual is None or actual == "" or actual == []:
                return False
        elif operator == "not_null" or operator == "is_not_null":
            if actual is None:
                return False
        elif operator == "is_null":
            if actual is not None:
                return False
        elif operator == "eq":
            if str(actual) != str(expected):
                return False
        elif operator == "neq":
            if str(actual) == str(expected):
                return False
        elif operator == "contains":
            if expected not in str(actual or ""):
                return False
        elif operator == "not_contains":
            if expected in str(actual or ""):
                return False
        elif operator in ("gt", "lt", "gte", "lte"):
            try:
                a_val = float(str(actual or 0))
                e_val = float(str(expected or 0))
                if operator == "gt" and not (a_val > e_val):
                    return False
                if operator == "lt" and not (a_val < e_val):
                    return False
                if operator == "gte" and not (a_val >= e_val):
                    return False
                if operator == "lte" and not (a_val <= e_val):
                    return False
            except (ValueError, TypeError):
                return False
    return True


def _apply_extract_company(title: str) -> str:
    title = str(title or "").strip()
    title = re.sub(
        r"^(demo|call|intro|sync|meeting|review|catch[- ]?up|discussion|"
        r"follow[- ]?up|check[- ]?in|onboarding|kickoff|discovery)\s+(with|for|from|@)?\s*",
        "", title, flags=re.IGNORECASE,
    ).strip()
    title = re.sub(r"\s*[-–|]\s*\d{4}[-/]\d{2}[-/]\d{2}.*$", "", title).strip()
    title = re.sub(
        r"\s*[-–|]\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s*\d+.*$",
        "", title, flags=re.IGNORECASE,
    ).strip()
    for sep in (" - ", " | ", " — "):
        if sep in title:
            parts = [p.strip() for p in title.split(sep)]
            title = parts[0] if parts[0] else parts[-1]
            break
    return title.strip()


_TITLE_ALIASES = ("meeting_title", "title", "name", "subject", "summary", "description")


def _resolve_field(record: dict, field: str) -> str:
    val = record.get(field)
    if val:
        return str(val)
    for alias in _TITLE_ALIASES:
        if alias != field and record.get(alias):
            return str(record[alias])
    return ""


"""
MAP NODE — TRANSFORM REFERENCE
===============================
Use inside the "mappings" object of a MAP node config.

SIMPLE RENAME (no transform)
─────────────────────────────
  "source_field": "target_field"
  Copies the value as-is under a new name.

  Example:
    "category": "TipoAlerta"

TRANSFORM OBJECT
─────────────────
  "source_field": { "target": "target_field", "transform": "transform_type", "args": {...} }
  Renames AND transforms. "args" is optional.

  Example:
    "created_at": { "target": "FechaAlerta", "transform": "extract_date" }

ONE SOURCE → MULTIPLE TARGETS (array)
──────────────────────────────────────
  "source_field": [
    "plain_target",
    { "target": "another_target", "transform": "extract_time" }
  ]
  Produces several output fields from the same input field.

  Example:
    "created_at": [
      { "target": "FechaAlerta", "transform": "extract_date" },
      { "target": "HoraAlerta",  "transform": "extract_time" }
    ]

─────────────────────────────────────────────────────────────────────────────
AVAILABLE TRANSFORMS
─────────────────────────────────────────────────────────────────────────────

DATE / TIME  (source must be ISO 8601: "2026-04-06T16:29:56Z" or "2026-04-06")
  extract_date      "2026-04-06T16:29:56Z" → "2026-04-06"
                    args: { "format": "%d/%m/%Y" }  optional strftime string

  extract_time      "2026-04-06T16:29:56Z" → "16:29:56"
                    args: { "format": "%H:%M" }

  extract_datetime  Reformat a full datetime.
                    args: { "format": "%d/%m/%Y %H:%M" }

  extract_year      → "2026"
  extract_month     → "04"
  extract_day       → "06"
  extract_hour      → "16"
  extract_minute    → "29"

TYPE CONVERSION
  to_string         Any value → string.  123 → "123"
  to_number         "3.14" or "3,14" → 3.14  (int when whole number)
  to_boolean        "true"/"1"/"yes"/"si"/"sí" → true, everything else → false

STRING OPERATIONS
  lowercase         "Hello World" → "hello world"
  uppercase         "hello world" → "HELLO WORLD"
  strip             "  hello  "  → "hello"  (removes whitespace)

  truncate          Cut string at N characters.
                    args: { "length": 100 }   default 255

  replace           Find and replace inside a string.
                    args: { "find": "foo", "replace": "bar" }

  substring         Slice a string by character position.
                    args: { "start": 0, "end": 10 }   (end optional)

  template          Build a string from multiple record fields.
                    args: { "template": "{name} ({id})" }
                    Uses Python .format() — reference any field in the record.

LIST / CSV
  split_csv         "a, b, c" → ["a", "b", "c"]
                    args: { "sep": "," }   default ","

  first_csv         "a, b, c" → "a"  (first item only)
                    args: { "sep": "," }

  join_list         ["a", "b", "c"] → "a, b, c"
                    args: { "sep": ", " }   default ", "

FALLBACK
  default_if_null   Return a fallback when the source value is null/empty.
                    args: { "value": "N/A" }

LEGACY
  extract_company   Strips common suffixes (LLC, Inc, S.A., etc.) from a company name.
  (no args)

─────────────────────────────────────────────────────────────────────────────
FULL EXAMPLE — Denuncias pipeline MAP config
─────────────────────────────────────────────────────────────────────────────
{
  "mappings": {
    "category":         "TipoAlerta",
    "location":         "AlertaCalle",
    "lat":              "Latitud",
    "lon":              "Longitud",
    "created_at": [
      { "target": "FechaAlerta", "transform": "extract_date" },
      { "target": "HoraAlerta",  "transform": "extract_time" }
    ],
    "id":               "complaint_id",
    "text":             "description",
    "status":           "status",
    "priority":         "priority",
    "source":           "source",
    "link":             "source_url",
    "twitter_username": "author",
    "is_live_event":    { "target": "is_live_event", "transform": "to_boolean" },
    "media_urls":       { "target": "media_urls",    "transform": "split_csv" }
  }
}
─────────────────────────────────────────────────────────────────────────────
"""


def _apply_transform(record: dict, source_field: str, target_field: str, transform_type: str, transform_args: dict | None = None) -> dict:
    record = dict(record)
    source_val = _resolve_field(record, source_field)
    args = transform_args or {}

    # ── Legacy ────────────────────────────────────────────────────────────────
    if transform_type == "extract_company":
        record[target_field] = _apply_extract_company(source_val)

    # ── Case ──────────────────────────────────────────────────────────────────
    elif transform_type == "lowercase":
        record[target_field] = str(source_val).lower() if source_val is not None else source_val

    elif transform_type == "uppercase":
        record[target_field] = str(source_val).upper() if source_val is not None else source_val

    elif transform_type == "strip":
        record[target_field] = str(source_val).strip() if source_val is not None else source_val

    # ── Type conversion ───────────────────────────────────────────────────────
    elif transform_type == "to_string":
        record[target_field] = str(source_val) if source_val is not None else None

    elif transform_type == "to_number":
        try:
            v = float(str(source_val).replace(",", "."))
            record[target_field] = int(v) if v == int(v) else v
        except (ValueError, TypeError):
            record[target_field] = None

    elif transform_type == "to_boolean":
        if isinstance(source_val, bool):
            record[target_field] = source_val
        else:
            record[target_field] = str(source_val).lower() in ("true", "1", "yes", "si", "sí")

    # ── Date / time extraction ────────────────────────────────────────────────
    elif transform_type in ("extract_date", "extract_time", "extract_year",
                            "extract_month", "extract_day", "extract_hour",
                            "extract_minute", "extract_datetime"):
        from datetime import datetime as _dt
        val = source_val
        parsed = None
        if isinstance(val, str) and val:
            for fmt in ("%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S.%fZ",
                        "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    parsed = _dt.strptime(val.replace("+00:00", "Z"), fmt)
                    break
                except ValueError:
                    continue
        if parsed:
            fmt_str = args.get("format")
            if transform_type == "extract_date":
                record[target_field] = parsed.strftime(fmt_str or "%Y-%m-%d")
            elif transform_type == "extract_time":
                record[target_field] = parsed.strftime(fmt_str or "%H:%M:%S")
            elif transform_type == "extract_year":
                record[target_field] = str(parsed.year)
            elif transform_type == "extract_month":
                record[target_field] = parsed.strftime("%m")
            elif transform_type == "extract_day":
                record[target_field] = parsed.strftime("%d")
            elif transform_type == "extract_hour":
                record[target_field] = parsed.strftime("%H")
            elif transform_type == "extract_minute":
                record[target_field] = parsed.strftime("%M")
            elif transform_type == "extract_datetime":
                record[target_field] = parsed.strftime(fmt_str or "%Y-%m-%dT%H:%M:%SZ")
        else:
            record[target_field] = source_val

    # ── String utilities ──────────────────────────────────────────────────────
    elif transform_type == "split_csv":
        sep = args.get("sep", ",")
        if isinstance(source_val, str):
            record[target_field] = [v.strip() for v in source_val.split(sep) if v.strip()]
        else:
            record[target_field] = source_val

    elif transform_type == "join_list":
        sep = args.get("sep", ", ")
        if isinstance(source_val, list):
            record[target_field] = sep.join(str(v) for v in source_val)
        else:
            record[target_field] = source_val

    elif transform_type == "truncate":
        n = int(args.get("length", 255))
        record[target_field] = str(source_val)[:n] if source_val is not None else source_val

    elif transform_type == "replace":
        find = args.get("find", "")
        repl = args.get("replace", "")
        record[target_field] = str(source_val).replace(find, repl) if source_val is not None else source_val

    elif transform_type == "substring":
        start = int(args.get("start", 0))
        end = args.get("end")
        s = str(source_val) if source_val is not None else ""
        record[target_field] = s[start:int(end)] if end is not None else s[start:]

    elif transform_type == "template":
        tmpl = args.get("template", str(source_val))
        try:
            record[target_field] = tmpl.format(**record)
        except (KeyError, ValueError):
            record[target_field] = tmpl

    elif transform_type == "default_if_null":
        default = args.get("value", "")
        record[target_field] = source_val if source_val is not None else default

    elif transform_type == "first_csv":
        sep = args.get("sep", ",")
        parts = str(source_val).split(sep) if source_val else []
        record[target_field] = parts[0].strip() if parts else None

    else:
        # Default: plain rename / copy
        record[target_field] = source_val

    return record


# ── Legacy: array_append (keep for existing pipelines) ───────────────────────

async def _execute_array_append(
    pipeline: "Pipeline",
    ot_id: str,
    array_field: str,
    merge_key: str,
) -> int:
    source_connector_id: str | None = None
    map_transforms: list[dict] = []

    for node in pipeline.nodes:
        if node.type == NodeType.SOURCE:
            source_connector_id = (
                node.config.get("connectorId")
                or node.config.get("connector_id")
                or node.connector_id
                or (pipeline.connector_ids[0] if pipeline.connector_ids else None)
            )
        elif node.type == NodeType.MAP:
            cfg = node.config
            jke = cfg.get("join_key_extraction")
            if jke:
                sf = jke.get("source_field", "")
                tf = jke.get("output_field", "__join_key__")
                tt = jke.get("transform", "")
            else:
                sf = cfg.get("sourceField") or cfg.get("source_field", "")
                tf = cfg.get("targetField") or cfg.get("target_field") or cfg.get("output_field", "__join_key__")
                tt = cfg.get("transformType") or cfg.get("transform_type") or cfg.get("transform", "")
            if sf:
                map_transforms.append({"source_field": sf, "target_field": tf, "transform_type": tt})

    if not source_connector_id:
        return 0

    try:
        async with httpx.AsyncClient(timeout=60) as client:
            r = await client.get(
                f"{CONNECTOR_API}/connectors/{source_connector_id}/schema",
                headers={"x-tenant-id": pipeline.tenant_id or "tenant-001"},
            )
            if not r.is_success:
                return 0
            raw_records: list[dict] = r.json().get("sample_rows", [])
    except Exception:
        return 0

    if not raw_records:
        return 0

    transformed_records: list[dict] = []
    for rec in raw_records:
        for t in map_transforms:
            rec = _apply_transform(rec, t["source_field"], t["target_field"], t["transform_type"])
        transformed_records.append(rec)

    if not map_transforms:
        for rec in transformed_records:
            rec["__join_key__"] = _resolve_field(rec, "meeting_title")

    try:
        async with httpx.AsyncClient(timeout=60) as client:
            resp = await client.post(
                f"{ONTOLOGY_API}/object-types/{ot_id}/records/array-append",
                json={
                    "array_field": array_field,
                    "merge_key": merge_key,
                    "join_key": "__join_key__",
                    "records": transformed_records,
                },
                headers={"x-tenant-id": pipeline.tenant_id or "tenant-001"},
            )
            if resp.is_success:
                return resp.json().get("appended", 0)
    except Exception:
        pass

    return 0


# ── Legacy: emit events by reading from ontology (fallback) ──────────────────

async def _emit_events_from_records(
    object_type_id: str,
    pipeline_id: str,
    connector_ids: list[str],
    case_id_field: str,
    activity_field: str,
    timestamp_field: str,
    tenant_id: str = "tenant-001",
) -> int:
    connector_id = connector_ids[0] if connector_ids else ""

    try:
        async with httpx.AsyncClient(timeout=60) as client:
            resp = await client.get(
                f"{ONTOLOGY_API}/object-types/{object_type_id}/records",
                headers={"x-tenant-id": tenant_id},
            )
            if not resp.is_success:
                return 0
            data = resp.json()
    except Exception:
        return 0

    records: list[dict] = data.get("records", [])
    if not records:
        return 0

    def _pick_timestamp(record: dict, preferred: str) -> str | None:
        if preferred and record.get(preferred):
            return str(record[preferred])
        for key in ("createdate", "hs_lastmodifieddate", "closedate", "created_at",
                    "updated_at", "timestamp", "date", "occurred_at"):
            if record.get(key):
                return str(record[key])
        for key, val in record.items():
            if val and any(t in key.lower() for t in ("date", "time", "_at", "stamp")):
                return str(val)
        return datetime.now(timezone.utc).isoformat()

    def _pick_value(record: dict, field: str, fallback: str) -> str:
        if field and record.get(field) is not None:
            return str(record[field])
        return fallback

    events = []
    for record in records:
        case_id = _pick_value(record, case_id_field, str(record.get("id", str(uuid4()))))
        activity = _pick_value(record, activity_field, "RECORD_SYNCED")
        ts = _pick_timestamp(record, timestamp_field)

        if ts and ts.isdigit():
            try:
                ms = int(ts)
                ts = datetime.fromtimestamp(ms / 1000, tz=timezone.utc).isoformat()
            except Exception:
                pass

        events.append({
            "id": str(uuid4()),
            "case_id": case_id,
            "activity": activity,
            "timestamp": ts or datetime.now(timezone.utc).isoformat(),
            "object_type_id": object_type_id,
            "object_id": case_id,
            "pipeline_id": pipeline_id,
            "connector_id": connector_id,
            "tenant_id": tenant_id,
            "attributes": {
                k: v for k, v in record.items()
                if k not in (case_id_field, activity_field, timestamp_field)
                and not isinstance(v, (list, dict))
            },
        })

    if not events:
        return 0

    written = 0
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            for i in range(0, len(events), 200):
                chunk = events[i:i + 200]
                resp = await client.post(
                    f"{EVENT_LOG_API}/events/batch",
                    json={"events": chunk},
                    headers={"x-tenant-id": tenant_id},
                )
                if resp.is_success:
                    written += len(chunk)
                else:
                    logger.error("_emit_events_from_records batch POST failed: %s %s", resp.status_code, resp.text[:500])
    except Exception as exc:
        logger.error("_emit_events_from_records error: %s", exc, exc_info=True)

    return written


# ── Pipeline execution event emission ────────────────────────────────────────

async def _emit_pipeline_event(
    pipeline_id: str,
    pipeline_name: str,
    activity: str,
    timestamp: str,
    rows_in: int,
    rows_out: int,
    status: str,
    error: str = "",
    tenant_id: str = "tenant-001",
) -> None:
    """Emit a pipeline-level event to the event log so run history is visible."""
    try:
        async with httpx.AsyncClient(timeout=5) as client:
            await client.post(
                f"{EVENT_LOG_API}/events",
                json={
                    "id": str(uuid4()),
                    "case_id": pipeline_id,
                    "activity": activity,
                    "timestamp": timestamp,
                    "object_type_id": "",
                    "object_id": pipeline_id,
                    "pipeline_id": pipeline_id,
                    "connector_id": "",
                    "tenant_id": tenant_id,
                    "attributes": {
                        "pipeline_name": pipeline_name,
                        "rows_in": rows_in,
                        "rows_out": rows_out,
                        "status": status,
                        "error": error,
                    },
                },
            )
    except Exception:
        pass
