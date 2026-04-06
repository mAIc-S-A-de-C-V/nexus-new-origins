import httpx
import io


async def run(inputs: dict) -> dict:
    file_url = inputs.get("file_url", "")
    sheet_name = inputs.get("sheet")
    header_row = int(inputs.get("header_row", 0))

    if not file_url:
        return {"error": "file_url is required"}

    async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
        try:
            resp = await client.get(file_url)
            if not resp.is_success:
                return {"error": f"Failed to download file: HTTP {resp.status_code}"}
            file_bytes = resp.content
        except Exception as e:
            return {"error": f"Download failed: {str(e)}"}

    url_lower = file_url.lower().split("?")[0]
    is_csv = url_lower.endswith(".csv")

    try:
        if is_csv:
            import pandas as pd
            df = pd.read_csv(io.BytesIO(file_bytes), header=header_row)
            headers = list(df.columns.astype(str))
            rows = df.fillna("").astype(str).to_dict(orient="records")
            sheets = [{"name": "Sheet1", "headers": headers, "rows": rows}]
            return {"sheets": sheets, "row_count": len(rows)}
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            sheet_names = wb.sheetnames

            if sheet_name and sheet_name in sheet_names:
                target_sheets = [sheet_name]
            elif sheet_name:
                try:
                    idx = int(sheet_name)
                    target_sheets = [sheet_names[idx]] if 0 <= idx < len(sheet_names) else [sheet_names[0]]
                except (ValueError, IndexError):
                    target_sheets = [sheet_names[0]]
            else:
                target_sheets = sheet_names

            result_sheets = []
            total_rows = 0
            for sname in target_sheets:
                ws = wb[sname]
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    result_sheets.append({"name": sname, "headers": [], "rows": []})
                    continue

                h_idx = header_row if header_row < len(all_rows) else 0
                headers = [str(c) if c is not None else "" for c in all_rows[h_idx]]
                data_rows = []
                for row in all_rows[h_idx + 1:]:
                    data_rows.append({
                        headers[i]: (str(v) if v is not None else "")
                        for i, v in enumerate(row)
                        if i < len(headers)
                    })
                total_rows += len(data_rows)
                result_sheets.append({"name": sname, "headers": headers, "rows": data_rows})

            return {"sheets": result_sheets, "row_count": total_rows}
    except Exception as e:
        return {"error": f"Parse failed: {str(e)}"}
