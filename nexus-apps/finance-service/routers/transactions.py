"""
Expense transaction ledger.
  GET    /finance/transactions           list (filter by category, year, month)
  POST   /finance/transactions           create
  PUT    /finance/transactions/{id}      update
  DELETE /finance/transactions/{id}      delete
  POST   /finance/transactions/upload    parse & import Excel file
  GET    /finance/transactions/summary   totals grouped by category + month
"""
import io
import uuid
from datetime import date
from decimal import Decimal
from typing import List, Optional

from fastapi import APIRouter, Depends, Header, HTTPException, UploadFile, File
from pydantic import BaseModel
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession

from database import TransactionRow, get_session

router = APIRouter()

TENANT = "tenant-001"
CATEGORIES = {"salaries", "software", "admin", "finanzas", "oficina", "marketing"}

# ── Pydantic schemas ──────────────────────────────────────────────────────────

class TransactionIn(BaseModel):
    category: str
    date: str
    description: str = ""
    vendor: Optional[str] = None
    payment_method: Optional[str] = None
    amount_usd: float
    notes: Optional[str] = None


class TransactionOut(BaseModel):
    id: str
    tenant_id: str
    category: str
    date: str
    description: str
    vendor: Optional[str]
    payment_method: Optional[str]
    amount_usd: float
    notes: Optional[str]
    created_at: Optional[str]

    class Config:
        from_attributes = True


def _row_to_out(r: TransactionRow) -> dict:
    return {
        "id": r.id,
        "tenant_id": r.tenant_id,
        "category": r.category,
        "date": r.date,
        "description": r.description or "",
        "vendor": r.vendor,
        "payment_method": r.payment_method,
        "amount_usd": float(r.amount_usd),
        "notes": r.notes,
        "created_at": r.created_at.isoformat() if r.created_at else None,
    }


def _tid(x_tenant_id: str = Header(default=TENANT)) -> str:
    return x_tenant_id


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("")
async def list_transactions(
    category: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    tenant_id: str = Depends(_tid),
    db: AsyncSession = Depends(get_session),
):
    q = select(TransactionRow).where(TransactionRow.tenant_id == tenant_id)
    if category:
        q = q.where(TransactionRow.category == category)
    if year:
        q = q.where(TransactionRow.date.like(f"{year}-%"))
    if month:
        m = str(month).zfill(2)
        q = q.where(TransactionRow.date.like(f"%-{m}-%"))
    q = q.order_by(TransactionRow.date.desc())
    rows = (await db.execute(q)).scalars().all()
    return [_row_to_out(r) for r in rows]


@router.post("")
async def create_transaction(
    body: TransactionIn,
    tenant_id: str = Depends(_tid),
    db: AsyncSession = Depends(get_session),
):
    if body.category not in CATEGORIES:
        raise HTTPException(400, f"category must be one of: {sorted(CATEGORIES)}")
    row = TransactionRow(
        id=str(uuid.uuid4()),
        tenant_id=tenant_id,
        **body.model_dump(),
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return _row_to_out(row)


@router.put("/{txn_id}")
async def update_transaction(
    txn_id: str,
    body: TransactionIn,
    tenant_id: str = Depends(_tid),
    db: AsyncSession = Depends(get_session),
):
    row = await db.get(TransactionRow, txn_id)
    if not row or row.tenant_id != tenant_id:
        raise HTTPException(404)
    for k, v in body.model_dump().items():
        setattr(row, k, v)
    await db.commit()
    await db.refresh(row)
    return _row_to_out(row)


@router.delete("/{txn_id}")
async def delete_transaction(
    txn_id: str,
    tenant_id: str = Depends(_tid),
    db: AsyncSession = Depends(get_session),
):
    row = await db.get(TransactionRow, txn_id)
    if not row or row.tenant_id != tenant_id:
        raise HTTPException(404)
    await db.delete(row)
    await db.commit()
    return {"deleted": txn_id}


# ── Excel upload ──────────────────────────────────────────────────────────────

_SHEET_CATEGORY = {
    "salaries":         "salaries",
    "software":         "software",
    "admin.":           "admin",
    "admin":            "admin",
    "finanzas":         "finanzas",
    "oficina":          "oficina",
    "office (oficina)": "oficina",
    "marketing. sales": "marketing",
    "marketing & sales":"marketing",
    "marketing":        "marketing",
}


def _parse_date(val) -> Optional[str]:
    """Convert Excel date cell (datetime or string) to YYYY-MM-DD."""
    if val is None:
        return None
    from datetime import datetime
    if isinstance(val, datetime):
        return val.date().isoformat()
    s = str(val).strip()
    # try DD/MM/YYYY
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


@router.post("/upload")
async def upload_excel(
    file: UploadFile = File(...),
    tenant_id: str = Depends(_tid),
    db: AsyncSession = Depends(get_session),
):
    import openpyxl

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    created = []
    skipped = 0

    for sheet_name in wb.sheetnames:
        cat_key = sheet_name.strip().lower()
        category = _SHEET_CATEGORY.get(cat_key)
        if not category:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue

        # Row 3 (index 2) is the header row: Date | Description | Vendor | PaymentMethod | Amount
        # Find column indices by scanning header row
        header = [str(c).strip().lower() if c else "" for c in rows[2]]

        def col(keywords):
            for kw in keywords:
                for i, h in enumerate(header):
                    if kw in h:
                        return i
            return None

        idx_date   = col(["date"])
        idx_desc   = col(["description", "roles", "desc"])
        idx_vendor = col(["vendor", "payee"])
        idx_pm     = col(["payment method", "payment"])
        idx_amount = col(["amount"])

        if idx_date is None or idx_amount is None:
            continue

        for row in rows[3:]:
            date_val   = row[idx_date]   if idx_date   is not None else None
            amount_val = row[idx_amount] if idx_amount is not None else None

            # Skip empty or total rows
            if amount_val is None:
                continue
            try:
                amount = float(str(amount_val).replace(",", "").strip())
            except (ValueError, TypeError):
                skipped += 1
                continue

            if amount <= 0:
                skipped += 1
                continue

            date_str = _parse_date(date_val)
            if not date_str:
                skipped += 1
                continue

            # Skip obvious total rows
            desc_raw = str(row[idx_desc]).strip().upper() if idx_desc is not None and row[idx_desc] else ""
            if "TOTAL" in desc_raw:
                skipped += 1
                continue

            txn = TransactionRow(
                id=str(uuid.uuid4()),
                tenant_id=tenant_id,
                category=category,
                date=date_str,
                description=str(row[idx_desc]).strip() if idx_desc is not None and row[idx_desc] else "",
                vendor=str(row[idx_vendor]).strip() if idx_vendor is not None and row[idx_vendor] else None,
                payment_method=str(row[idx_pm]).strip() if idx_pm is not None and row[idx_pm] else None,
                amount_usd=amount,
            )
            db.add(txn)
            created.append(txn.id)

    await db.commit()
    return {"imported": len(created), "skipped": skipped}


# ── Summary ───────────────────────────────────────────────────────────────────

@router.get("/summary")
async def summary(
    year: Optional[int] = None,
    tenant_id: str = Depends(_tid),
    db: AsyncSession = Depends(get_session),
):
    """Return totals grouped by category and by month (YYYY-MM)."""
    q = select(TransactionRow).where(TransactionRow.tenant_id == tenant_id)
    if year:
        q = q.where(TransactionRow.date.like(f"{year}-%"))
    rows = (await db.execute(q)).scalars().all()

    by_category: dict[str, float] = {}
    by_month: dict[str, float] = {}

    for r in rows:
        amt = float(r.amount_usd)
        by_category[r.category] = by_category.get(r.category, 0) + amt
        month = r.date[:7]  # YYYY-MM
        by_month[month] = by_month.get(month, 0) + amt

    return {
        "total": sum(by_category.values()),
        "by_category": by_category,
        "by_month": dict(sorted(by_month.items())),
    }
